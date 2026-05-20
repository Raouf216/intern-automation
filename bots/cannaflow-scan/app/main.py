import json
import os
import re
import time
import traceback
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from urllib.parse import quote

from fastapi import FastAPI, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse
from playwright.sync_api import expect
from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
from playwright.sync_api import sync_playwright


STARTED_AT = time.time()
SERVICE_NAME = "cannaflow-scan"
ARTIFACTS_DIR = os.environ.get("ARTIFACTS_DIR", "/app/artifacts")
SESSION_STATE_PATH = os.environ.get(
    "CANNAFLOW_SESSION_STATE_PATH",
    os.path.join(ARTIFACTS_DIR, "cannaflow-storage-state.json"),
)
DEFAULT_USER_AGENT = (
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/136.0.0.0 Safari/537.36"
)
SCRAPE_PAGE_READY_TIMEOUT_MS = int(os.environ.get("CANNAFLOW_SCRAPE_PAGE_READY_TIMEOUT_MS", "30000"))
SCRAPE_MAX_PAGES = int(os.environ.get("CANNAFLOW_SCRAPE_MAX_PAGES", "50"))
SUPABASE_TIMEOUT_SECONDS = int(os.environ.get("CANNAFLOW_SUPABASE_TIMEOUT_SECONDS", "60"))

app = FastAPI(title=SERVICE_NAME)

cors_origins = [
    origin.strip()
    for origin in os.environ.get(
        "CANNAFLOW_SCAN_CORS_ORIGINS",
        "http://localhost:8040,http://127.0.0.1:8040,http://178.104.144.30:8040",
    ).split(",")
    if origin.strip()
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=cors_origins,
    allow_methods=["GET", "POST", "OPTIONS"],
    allow_headers=["*"],
)


class BotStepError(RuntimeError):
    def __init__(self, message, **details):
        super().__init__(message)
        self.details = details


def bool_env(name, default=True):
    value = os.environ.get(name)
    if value is None:
        return default
    return value.strip().lower() in {"1", "true", "yes", "on"}


def int_env(name, default):
    value = os.environ.get(name)
    if value is None or not value.strip():
        return default
    try:
        return int(value)
    except ValueError as exc:
        raise RuntimeError(f"Invalid integer environment variable {name}={value!r}") from exc


def required_env(name):
    value = (os.environ.get(name) or "").strip()
    if not value:
        raise RuntimeError(f"Missing required environment variable: {name}")
    return value


def login_url():
    return (os.environ.get("CANNAFLOW_LOGIN_URL") or "").strip() or "https://erp.cannaflow.de/auth/login"


def inventory_url():
    return (os.environ.get("CANNAFLOW_INVENTORY_URL") or "").strip() or "https://erp.cannaflow.de/inventory"


def ready_text():
    return (os.environ.get("CANNAFLOW_READY_TEXT") or "Inventar").strip() or "Inventar"


def inventory_page_size():
    return int_env("CANNAFLOW_PAGE_SIZE", 200)


def products_schema():
    return (
        os.environ.get("CANNAFLOW_PRODUCTS_SCHEMA")
        or os.environ.get("SUPABASE_SCHEMA")
        or "private"
    ).strip() or "private"


def products_table():
    return (os.environ.get("CANNAFLOW_PRODUCTS_TABLE") or "cannaflow_products").strip() or "cannaflow_products"


def products_replace_all():
    return bool_env("CANNAFLOW_PRODUCTS_REPLACE_ALL", True)


def products_upsert_on():
    value = (os.environ.get("CANNAFLOW_PRODUCTS_UPSERT_ON") or "product_name").strip()
    return value or None


def stock_report_enabled():
    return bool_env("CANNAFLOW_STOCK_REPORT_ENABLED", True)


def supabase_url():
    return (os.environ.get("SUPABASE_URL") or "").strip().rstrip("/")


def supabase_service_role_key():
    return (os.environ.get("SUPABASE_SERVICE_ROLE_KEY") or "").strip()


def supabase_rest_configured():
    return bool(supabase_url() and supabase_service_role_key())


def user_agent():
    return (os.environ.get("CANNAFLOW_USER_AGENT") or DEFAULT_USER_AGENT).strip() or DEFAULT_USER_AGENT


def browser_context_options():
    return {
        "user_agent": user_agent(),
        "viewport": {"width": 1365, "height": 900},
        "locale": "de-DE",
        "timezone_id": os.environ.get("TZ", "Europe/Berlin"),
    }


def utc_now_iso():
    return datetime.now(timezone.utc).isoformat()


def base_url_from_request(request):
    return str(request.base_url).rstrip("/")


def safe_screenshot_filename(filename):
    if not re.match(r"^[A-Za-z0-9_.-]+\.png$", filename or ""):
        raise RuntimeError("Invalid screenshot filename.")
    return filename


def latest_screenshot_path():
    if not os.path.isdir(ARTIFACTS_DIR):
        return None

    screenshot_paths = [
        os.path.join(ARTIFACTS_DIR, filename)
        for filename in os.listdir(ARTIFACTS_DIR)
        if filename.lower().endswith(".png")
    ]
    if not screenshot_paths:
        return None
    return max(screenshot_paths, key=os.path.getmtime)


def safe_report_filename(filename):
    if not re.match(r"^[A-Za-z0-9_.-]+\.xlsx$", filename or ""):
        raise RuntimeError("Invalid report filename.")
    return filename


def latest_report_path():
    if not os.path.isdir(ARTIFACTS_DIR):
        return None

    report_paths = [
        os.path.join(ARTIFACTS_DIR, filename)
        for filename in os.listdir(ARTIFACTS_DIR)
        if filename.lower().endswith(".xlsx")
    ]
    if not report_paths:
        return None
    return max(report_paths, key=os.path.getmtime)


def screenshot_file_url(base_url, screenshot_path):
    return f"{base_url}/screenshots/{os.path.basename(screenshot_path)}"


def screenshot_response_links(base_url, screenshot_path):
    return {
        "screenshot_url": screenshot_file_url(base_url, screenshot_path),
        "latest_screenshot_url": f"{base_url}/screenshots/latest",
    }


def report_file_url(base_url, report_path):
    return f"{base_url}/reports/{os.path.basename(report_path)}"


def report_response_links(base_url, report_path):
    return {
        "report_url": report_file_url(base_url, report_path),
        "latest_report_url": f"{base_url}/reports/latest",
    }


def supabase_table_url():
    return f"{supabase_url()}/rest/v1/{quote(products_table(), safe='')}"


def supabase_headers():
    service_role_key = required_env("SUPABASE_SERVICE_ROLE_KEY")
    schema_name = products_schema()

    return {
        "apikey": service_role_key,
        "Authorization": f"Bearer {service_role_key}",
        "Accept-Profile": schema_name,
        "Content-Profile": schema_name,
        "Content-Type": "application/json",
    }


def response_preview(response):
    body = response.text
    if len(body) > 1000:
        body = f"{body[:1000]}..."

    return {
        "status_code": response.status_code,
        "ok": response.status_code < 400,
        "body": body,
    }


def missing_column_from_postgrest_response(response):
    try:
        payload = response.json()
    except Exception:
        payload = {}

    message = str(payload.get("message") or response.text or "")
    match = re.search(r"Could not find the '([^']+)' column", message)

    if match:
        return match.group(1)

    match = re.search(r"column\s+\w+\.([A-Za-z_][A-Za-z0-9_]*)\s+does not exist", message)

    if match:
        return match.group(1)

    return None


def validate_identifier(value, label):
    if not re.fullmatch(r"[A-Za-z_][A-Za-z0-9_]*", value or ""):
        raise RuntimeError(f"Invalid {label}: {value!r}")

    return value


def capture_debug_screenshot(page, reason, trace=None):
    os.makedirs(ARTIFACTS_DIR, exist_ok=True)
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    safe_reason = re.sub(r"[^A-Za-z0-9_.-]+", "-", reason).strip("-") or "debug"
    path = os.path.join(ARTIFACTS_DIR, f"cannaflow-{safe_reason}-{timestamp}.png")

    try:
        page.screenshot(path=path, full_page=True)
        trace_step(trace, "capture_debug_screenshot", path=path, reason=reason)
        return path
    except Exception as exc:
        trace_step(trace, "capture_debug_screenshot_failed", reason=reason, error=f"{type(exc).__name__}: {exc}")
        return None


def page_text_excerpt(page, limit=500):
    try:
        text = page.locator("body").inner_text(timeout=2_000)
    except Exception:
        return ""
    return re.sub(r"\s+", " ", text or "").strip()[:limit]


def trace_step(trace, name, **fields):
    if trace:
        trace(name, **fields)


def wait_for_inventory_ready(page, trace=None):
    timeout_ms = int_env("CANNAFLOW_READY_TIMEOUT_MS", 60_000)
    text = ready_text()
    trace_step(trace, "wait_for_inventory_ready", ready_text=text, timeout_ms=timeout_ms)

    try:
        page.get_by_text(text, exact=False).first.wait_for(state="visible", timeout=timeout_ms)
    except Exception as exc:
        screenshot_path = capture_debug_screenshot(page, "inventory-not-ready", trace=trace)
        details = {
            "current_url": page.url,
            "page_title": page.title(),
            "ready_text": text,
            "body_excerpt": page_text_excerpt(page, limit=800),
            "debug_screenshot_path": screenshot_path,
        }
        raise BotStepError("Inventory page did not become ready.", **details) from exc

    return {
        "ready_text": text,
        "body_excerpt": page_text_excerpt(page),
    }


def page_size_input(page):
    return page.locator("xpath=//*[normalize-space()='Seitengröße']/following::input[1]").first


def wait_for_inventory_page_size_applied(page, target_size, trace=None):
    timeout_ms = int_env("CANNAFLOW_PAGE_SIZE_TIMEOUT_MS", 30_000)
    trace_step(trace, "wait_for_inventory_page_size_applied", target_size=target_size, timeout_ms=timeout_ms)

    try:
        return page.wait_for_function(
            """
            (targetSize) => {
              const normalize = (value) => (value || '').replace(/\\s+/g, ' ').trim();
              const bodyText = normalize(document.body ? document.body.innerText : '');
              const ranges = Array.from(bodyText.matchAll(/(\\d+)\\s*-\\s*(\\d+)\\s+von\\s+(\\d+)/g));
              const rangeMatch = ranges.length ? ranges[ranges.length - 1] : null;

              if (!rangeMatch) {
                return false;
              }

              const from = Number(rangeMatch[1]);
              const to = Number(rangeMatch[2]);
              const total = Number(rangeMatch[3]);
              const expectedRows = Math.max(0, to - from + 1);
              const tables = Array.from(document.querySelectorAll('table'));
              const tableInfo = tables
                .map((candidate) => ({
                  element: candidate,
                  rows: Array.from(candidate.querySelectorAll('tbody tr'))
                    .filter((row) => row.querySelectorAll('td').length > 0),
                }))
                .sort((left, right) => right.rows.length - left.rows.length)[0];
              const rowCount = tableInfo ? tableInfo.rows.length : 0;
              const requiredRows = Math.min(expectedRows, Number(targetSize) || expectedRows);

              if (!requiredRows || rowCount < requiredRows) {
                window.__cannaflowPageSizeProbe = {
                  label: rangeMatch[0],
                  expectedRows,
                  requiredRows,
                  rowCount,
                  total,
                };
                return false;
              }

              return {
                label: rangeMatch[0],
                expected_rows: expectedRows,
                row_count: rowCount,
                total_rows: total,
              };
            }
            """,
            arg=target_size,
            timeout=timeout_ms,
        ).json_value()
    except Exception as exc:
        probe = page.evaluate("() => window.__cannaflowPageSizeProbe || null")
        screenshot_path = capture_debug_screenshot(page, "page-size-table-not-ready", trace=trace)
        raise BotStepError(
            "Cannaflow table did not render the selected page size.",
            current_url=page.url,
            page_title=page.title(),
            page_size_target=target_size,
            page_size_probe=probe,
            body_excerpt=page_text_excerpt(page, limit=800),
            debug_screenshot_path=screenshot_path,
        ) from exc


def ensure_inventory_page_size(page, trace=None):
    target_size = inventory_page_size()
    timeout_ms = int_env("CANNAFLOW_PAGE_SIZE_TIMEOUT_MS", 30_000)
    page_size = page_size_input(page)

    trace_step(trace, "ensure_inventory_page_size", target_size=target_size, timeout_ms=timeout_ms)

    try:
        page_size.wait_for(state="visible", timeout=timeout_ms)
        current_value = (page_size.input_value(timeout=timeout_ms) or "").strip()
    except Exception as exc:
        screenshot_path = capture_debug_screenshot(page, "page-size-input-not-found", trace=trace)
        raise BotStepError(
            "Cannaflow page size input was not visible.",
            current_url=page.url,
            page_title=page.title(),
            body_excerpt=page_text_excerpt(page, limit=800),
            debug_screenshot_path=screenshot_path,
        ) from exc

    trace_step(trace, "inventory_page_size_current", current_value=current_value, target_size=target_size)

    try:
        current_size = int(current_value)
    except ValueError:
        current_size = 0

    if current_size >= target_size:
        render_state = wait_for_inventory_page_size_applied(page, target_size, trace=trace)
        return {
            "target": target_size,
            "before": current_value,
            "after": current_value,
            "changed": False,
            "render": render_state,
        }

    try:
        page_size.scroll_into_view_if_needed(timeout=timeout_ms)
        page_size.click(timeout=timeout_ms)
        page.get_by_role("option", name=str(target_size), exact=True).click(timeout=timeout_ms)
        expect(page_size).to_have_value(str(target_size), timeout=timeout_ms)
    except Exception as exc:
        screenshot_path = capture_debug_screenshot(page, "page-size-select-failed", trace=trace)
        raise BotStepError(
            "Could not set Cannaflow page size.",
            current_url=page.url,
            page_title=page.title(),
            page_size_before=current_value,
            page_size_target=target_size,
            body_excerpt=page_text_excerpt(page, limit=800),
            debug_screenshot_path=screenshot_path,
        ) from exc

    render_state = wait_for_inventory_page_size_applied(page, target_size, trace=trace)
    after_value = (page_size.input_value(timeout=timeout_ms) or "").strip()
    trace_step(trace, "inventory_page_size_changed", before=current_value, after=after_value)

    return {
        "target": target_size,
        "before": current_value,
        "after": after_value,
        "changed": True,
        "render": render_state,
    }


def normalize_space(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def parse_decimal(value):
    text = normalize_space(value)

    if not text or text.lower() in {"n.a.", "na", "n/a", "-"}:
        return None

    text = text.replace("€", "").replace("g", "").replace("/", "").replace(" ", "")
    text = text.replace(".", "").replace(",", ".")
    match = re.search(r"-?\d+(?:\.\d+)?", text)

    if not match:
        return None

    try:
        return Decimal(match.group(0))
    except InvalidOperation:
        return None


def parse_date(value):
    text = normalize_space(value)

    if not text or text.lower() in {"n.a.", "na", "n/a", "-"}:
        return None

    match = re.fullmatch(r"(\d{1,2})\.(\d{1,2})\.(\d{4})", text)
    if not match:
        return None

    day, month, year = map(int, match.groups())

    try:
        return datetime(year, month, day).date()
    except ValueError:
        return None


def json_ready(value):
    if isinstance(value, Decimal):
        return int(value) if value == value.to_integral_value() else float(value)

    if isinstance(value, datetime):
        return value.isoformat()

    if hasattr(value, "isoformat") and not isinstance(value, str):
        return value.isoformat()

    if isinstance(value, list):
        return [json_ready(item) for item in value]

    if isinstance(value, dict):
        return {key: json_ready(item) for key, item in value.items()}

    return value


def chunks(items, size):
    for index in range(0, len(items), size):
        yield items[index : index + size]


def decimal_for_compare(value):
    if value is None:
        return None

    if isinstance(value, Decimal):
        return value

    if isinstance(value, bool):
        return None

    if isinstance(value, (int, float)):
        try:
            return Decimal(str(value))
        except InvalidOperation:
            return None

    return parse_decimal(value)


def quantity_display(value, fallback=None):
    parsed = decimal_for_compare(value)

    if parsed is None:
        return normalize_space(fallback if fallback is not None else value)

    if parsed == parsed.to_integral_value():
        return str(int(parsed))

    return str(parsed).replace(".", ",")


def excel_number(value):
    parsed = decimal_for_compare(value)

    if parsed is None:
        return None

    if parsed == parsed.to_integral_value():
        return int(parsed)

    return float(parsed)


def page_says_not_authenticated(page):
    return "Nicht angemeldet" in page_text_excerpt(page, limit=1_000)


def click_inventory(page, trace=None):
    trace_step(trace, "open_inventory", url=inventory_url())

    inventory_link = page.get_by_role("link", name="Inventar")
    try:
        if inventory_link.count() > 0 and inventory_link.first.is_visible(timeout=2_000):
            trace_step(trace, "click_inventory_link")
            inventory_link.first.click(timeout=10_000)
        else:
            trace_step(trace, "goto_inventory_url")
            page.goto(inventory_url(), wait_until="domcontentloaded", timeout=60_000)
    except Exception:
        trace_step(trace, "inventory_link_failed_goto_url")
        page.goto(inventory_url(), wait_until="domcontentloaded", timeout=60_000)

    return wait_for_inventory_ready(page, trace=trace)


def login_form_visible(page):
    try:
        email = page.get_by_role("textbox", name="E-Mail", exact=True)
        password = page.get_by_role("textbox", name="Passwort", exact=True)
        return email.count() > 0 and password.count() > 0 and email.first.is_visible(timeout=2_000)
    except Exception:
        return False


def fill_login_form(page, trace=None):
    username = required_env("CANNAFLOW_USERNAME")
    password = required_env("CANNAFLOW_PASSWORD")

    trace_step(trace, "fill_login_form")
    page.get_by_role("textbox", name="E-Mail", exact=True).first.fill(username, timeout=15_000)
    page.get_by_role("textbox", name="Passwort", exact=True).first.fill(password, timeout=15_000)


def submit_login_form(page, trace=None):
    trace_step(trace, "submit_login_form")
    page.get_by_role("button", name="Anmelden", exact=True).click(timeout=15_000)


def wait_for_authenticated_shell(page, trace=None):
    timeout_ms = int_env("CANNAFLOW_POST_LOGIN_TIMEOUT_MS", 30_000)
    deadline = time.time() + (timeout_ms / 1000)
    trace_step(trace, "wait_for_authenticated_shell", timeout_ms=timeout_ms)

    while time.time() < deadline:
        if page_says_not_authenticated(page):
            break

        try:
            inventory_link = page.get_by_role("link", name="Inventar")
            if inventory_link.count() > 0 and inventory_link.first.is_visible(timeout=1_000):
                trace_step(trace, "authenticated_shell_visible", current_url=page.url)
                return
        except Exception:
            pass

        if "/auth/login" not in page.url:
            try:
                page.get_by_text("Inventar", exact=False).first.wait_for(state="visible", timeout=1_000)
                trace_step(trace, "authenticated_text_visible", current_url=page.url)
                return
            except Exception:
                pass

        page.wait_for_timeout(500)

    screenshot_path = capture_debug_screenshot(page, "post-login-not-authenticated", trace=trace)
    raise BotStepError(
        "Login did not reach the authenticated Cannaflow app.",
        current_url=page.url,
        page_title=page.title(),
        body_excerpt=page_text_excerpt(page, limit=800),
        debug_screenshot_path=screenshot_path,
    )


def open_context_with_saved_session(browser, trace=None):
    if not os.path.exists(SESSION_STATE_PATH):
        trace_step(trace, "saved_session_missing", path=SESSION_STATE_PATH)
        return None, None, False

    trace_step(trace, "open_saved_session", path=SESSION_STATE_PATH)
    context = browser.new_context(storage_state=SESSION_STATE_PATH, **browser_context_options())
    page = context.new_page()

    try:
        page.goto(inventory_url(), wait_until="domcontentloaded", timeout=60_000)
        wait_for_inventory_ready(page, trace=trace)
        return context, page, True
    except Exception as exc:
        trace_step(trace, "saved_session_not_ready", error=f"{type(exc).__name__}: {exc}", current_url=page.url)
        context.close()
        return None, None, False


def open_fresh_session(browser, trace=None):
    trace_step(trace, "open_fresh_session")
    context = browser.new_context(**browser_context_options())
    page = context.new_page()

    trace_step(trace, "goto_login_url", url=login_url())
    page.goto(login_url(), wait_until="domcontentloaded", timeout=60_000)
    trace_step(
        trace,
        "login_page_loaded",
        current_url=page.url,
        page_title=page.title(),
        body_excerpt=page_text_excerpt(page),
    )

    if not login_form_visible(page):
        trace_step(trace, "login_form_not_visible_try_inventory")
        page.goto(inventory_url(), wait_until="domcontentloaded", timeout=60_000)
    else:
        fill_login_form(page, trace=trace)
        submit_login_form(page, trace=trace)
        wait_for_authenticated_shell(page, trace=trace)

    click_inventory(page, trace=trace)
    os.makedirs(os.path.dirname(SESSION_STATE_PATH), exist_ok=True)
    context.storage_state(path=SESSION_STATE_PATH)
    trace_step(trace, "saved_session", path=SESSION_STATE_PATH)

    return context, page, False


def open_authenticated_inventory(browser, trace=None):
    context, page, reused_session = open_context_with_saved_session(browser, trace=trace)
    if context and page:
        return context, page, reused_session

    return open_fresh_session(browser, trace=trace)


def wait_for_inventory_rows_ready(page, trace=None):
    trace_step(trace, "wait_for_inventory_rows_ready", timeout_ms=SCRAPE_PAGE_READY_TIMEOUT_MS)
    page.wait_for_function(
        """
        () => {
          const normalize = (value) => (value || '').replace(/\\s+/g, ' ').trim();
          const tables = Array.from(document.querySelectorAll('table'));
          const table = tables
            .map((candidate) => ({
              element: candidate,
              rows: Array.from(candidate.querySelectorAll('tbody tr'))
                .filter((row) => row.querySelectorAll('td').length > 0),
            }))
            .sort((left, right) => right.rows.length - left.rows.length)[0];

          if (!table || table.rows.length === 0) {
            window.__cannaflowRowsReadyProbe = null;
            return false;
          }

          const first = normalize(table.rows[0].innerText);
          const last = normalize(table.rows[table.rows.length - 1].innerText);
          const key = `${table.rows.length}|${first}|${last}`;
          const previous = window.__cannaflowRowsReadyProbe || {};

          if (previous.key === key) {
            previous.hits = (previous.hits || 1) + 1;
            window.__cannaflowRowsReadyProbe = previous;
          } else {
            window.__cannaflowRowsReadyProbe = { key, hits: 1 };
          }

          return window.__cannaflowRowsReadyProbe.hits >= 2;
        }
        """,
        timeout=SCRAPE_PAGE_READY_TIMEOUT_MS,
    )


def scrape_current_inventory_page(page, page_number=None, trace=None):
    wait_for_inventory_rows_ready(page, trace=trace)
    result = page.evaluate(
        """
        () => {
          const normalize = (value) => (value || '').replace(/\\s+/g, ' ').trim();
          const fold = (value) => normalize(value)
            .normalize('NFD')
            .replace(/[\\u0300-\\u036f]/g, '')
            .toLowerCase();
          const slug = (value) => fold(value)
            .replace(/[^a-z0-9]+/g, '_')
            .replace(/^_+|_+$/g, '');
          const cleanText = (element) => {
            if (!element) return '';
            const clone = element.cloneNode(true);
            clone.querySelectorAll(
              'button, svg, input, .mantine-Checkbox-root, .mantine-Switch-root, [role="switch"]'
            ).forEach((node) => node.remove());
            return normalize(clone.textContent);
          };
          const switchValue = (cell) => {
            const roleSwitch = cell.querySelector('[role="switch"]');
            if (roleSwitch) {
              const checked = roleSwitch.getAttribute('aria-checked');
              if (checked === 'true') return true;
              if (checked === 'false') return false;
            }

            const switchInput = cell.querySelector('input[type="checkbox"]');
            if (!switchInput) return null;
            return Boolean(switchInput.checked);
          };
          const mapHeader = (header, index) => {
            const key = fold(header);
            if (key === 'name' || key.includes('produkt')) return 'product_name';
            if (key.includes('verfugbar')) return 'available_text';
            if (key === 'bestand' || key.includes('bestand')) return 'stock_text';
            if (key.includes('vk') || key.includes('preis')) return 'sale_price_text';
            if (key === 'art') return 'product_type';
            if (key.includes('kultivar')) return 'cultivar';
            if (key.includes('hersteller')) return 'manufacturer';
            if (key.includes('verkauf')) return 'sale_enabled';
            if (key.includes('verfallsdatum') || key.includes('verfall')) return 'expiry_date_text';
            return slug(header) || `column_${index + 1}`;
          };

          const tableInfo = Array.from(document.querySelectorAll('table'))
            .map((candidate) => ({
              element: candidate,
              rows: Array.from(candidate.querySelectorAll('tbody tr'))
                .filter((row) => row.querySelectorAll('td').length > 0),
            }))
            .sort((left, right) => right.rows.length - left.rows.length)[0];

          if (!tableInfo || !tableInfo.element || tableInfo.rows.length === 0) {
            return { ok: false, error: 'inventory_table_not_found' };
          }

          const table = tableInfo.element;
          let headers = Array.from(table.querySelectorAll('thead th')).map((header, index) => ({
            label: cleanText(header) || `column_${index + 1}`,
            field: mapHeader(cleanText(header), index),
          }));

          if (headers.length === 0) {
            headers = [
              'Name',
              'Verfügbar',
              'Bestand',
              'VK-Preis (brutto)',
              'Art',
              'Kultivar',
              'Hersteller',
              'Verkauf',
              'Verfallsdatum',
            ].map((label, index) => ({ label, field: mapHeader(label, index) }));
          }

          const rows = tableInfo.rows.map((row, rowIndex) => {
            const record = {
              row_index: rowIndex + 1,
              raw_cells: [],
            };

            Array.from(row.querySelectorAll('td')).forEach((cell, cellIndex) => {
              const header = headers[cellIndex] || {
                label: `column_${cellIndex + 1}`,
                field: `column_${cellIndex + 1}`,
              };
              const text = cleanText(cell);
              const toggle = switchValue(cell);

              record.raw_cells.push({
                index: cellIndex + 1,
                header: header.label,
                field: header.field,
                text,
                switch_value: toggle,
              });

              if (header.field === 'sale_enabled') {
                record.sale_enabled = toggle;
                return;
              }

              record[header.field] = text;
            });

            return record;
          });

          return {
            ok: true,
            headers,
            row_count: rows.length,
            rows,
          };
        }
        """
    )

    if not result.get("ok"):
        raise RuntimeError(result.get("error") or "inventory_scrape_failed")

    trace_step(
        trace,
        "scraped_inventory_page",
        page_number=page_number,
        row_count=result.get("row_count", 0),
    )
    return result


def get_pagination_state(page):
    return page.evaluate(
        """
        () => {
          const normalize = (value) => (value || '').replace(/\\s+/g, ' ').trim();
          const bodyText = normalize(document.body ? document.body.innerText : '');
          const ranges = Array.from(bodyText.matchAll(/(\\d+)\\s*-\\s*(\\d+)\\s+von\\s+(\\d+)/g));
          const rangeMatch = ranges.length ? ranges[ranges.length - 1] : null;
          const nextButton = Array.from(document.querySelectorAll('button'))
            .find((button) => normalize(button.innerText) === 'Weiter');
          const hasNext = Boolean(nextButton) && !(
            nextButton.disabled ||
            nextButton.getAttribute('aria-disabled') === 'true' ||
            nextButton.dataset.disabled === 'true'
          );
          const table = Array.from(document.querySelectorAll('table'))
            .map((candidate) => ({
              element: candidate,
              rows: Array.from(candidate.querySelectorAll('tbody tr'))
                .filter((row) => row.querySelectorAll('td').length > 0),
            }))
            .sort((left, right) => right.rows.length - left.rows.length)[0];
          const rows = table ? table.rows : [];
          const signature = rows.map((row) => normalize(row.innerText)).join('||');

          return {
            label: rangeMatch ? rangeMatch[0] : '',
            from_row: rangeMatch ? Number(rangeMatch[1]) : null,
            to_row: rangeMatch ? Number(rangeMatch[2]) : null,
            total_rows: rangeMatch ? Number(rangeMatch[3]) : null,
            has_next: hasNext,
            visible_row_count: rows.length,
            visible_product_signature: signature,
          };
        }
        """
    )


def click_next_inventory_page(page, before_state, trace=None):
    trace_step(trace, "click_next_inventory_page", before_state=before_state)
    next_button = page.get_by_role("button", name="Weiter", exact=True)

    if next_button.count() == 0:
        return {"clicked": False, "reason": "next_button_not_found"}

    if next_button.first.is_disabled(timeout=1_000):
        return {"clicked": False, "reason": "next_button_disabled"}

    next_button.first.click(timeout=SCRAPE_PAGE_READY_TIMEOUT_MS)
    page.wait_for_function(
        """
        (before) => {
          const normalize = (value) => (value || '').replace(/\\s+/g, ' ').trim();
          const bodyText = normalize(document.body ? document.body.innerText : '');
          const ranges = Array.from(bodyText.matchAll(/(\\d+)\\s*-\\s*(\\d+)\\s+von\\s+(\\d+)/g));
          const rangeMatch = ranges.length ? ranges[ranges.length - 1] : null;
          const currentFrom = rangeMatch ? Number(rangeMatch[1]) : null;
          const table = Array.from(document.querySelectorAll('table'))
            .map((candidate) => ({
              element: candidate,
              rows: Array.from(candidate.querySelectorAll('tbody tr'))
                .filter((row) => row.querySelectorAll('td').length > 0),
            }))
            .sort((left, right) => right.rows.length - left.rows.length)[0];
          const rows = table ? table.rows : [];
          const signature = rows.map((row) => normalize(row.innerText)).join('||');

          return (
            rows.length > 0 &&
            signature &&
            signature !== before.visible_product_signature &&
            (!before.from_row || !currentFrom || currentFrom > before.from_row)
          );
        }
        """,
        arg=before_state,
        timeout=SCRAPE_PAGE_READY_TIMEOUT_MS,
    )
    wait_for_inventory_rows_ready(page, trace=trace)
    return {"clicked": True}


def normalize_scraped_product(raw_product, page_number, source_url, scraped_at):
    available_text = normalize_space(raw_product.get("available_text"))
    stock_text = normalize_space(raw_product.get("stock_text"))
    sale_price_text = normalize_space(raw_product.get("sale_price_text"))
    expiry_date_text = normalize_space(raw_product.get("expiry_date_text"))
    raw_product = {
        **raw_product,
        "page_number": page_number,
        "source_url": source_url,
        "scraped_at": scraped_at,
    }

    return {
        "product_name": normalize_space(raw_product.get("product_name")),
        "available_text": available_text,
        "available_grams": parse_decimal(available_text),
        "stock_text": stock_text,
        "stock_grams": parse_decimal(stock_text),
        "sale_price_text": sale_price_text,
        "sale_price_per_g": parse_decimal(sale_price_text),
        "product_type": normalize_space(raw_product.get("product_type")),
        "cultivar": normalize_space(raw_product.get("cultivar")),
        "manufacturer": normalize_space(raw_product.get("manufacturer")),
        "sale_enabled": raw_product.get("sale_enabled"),
        "expiry_date_text": expiry_date_text,
        "expiry_date": parse_date(expiry_date_text),
        "page_number": page_number,
        "row_index": raw_product.get("row_index"),
        "source_url": source_url,
        "scraped_at": scraped_at,
        "raw_data": raw_product,
    }


def product_name_key(product):
    return normalize_space(product.get("product_name")).casefold()


def dedupe_products_by_name(products):
    deduped = []
    seen = set()
    duplicate_samples = []
    duplicate_count = 0

    for product in products:
        key = product_name_key(product)
        if not key:
            continue

        if key in seen:
            duplicate_count += 1
            if len(duplicate_samples) < 10:
                duplicate_samples.append(product.get("product_name"))
            continue

        seen.add(key)
        deduped.append(product)

    return deduped, {
        "input_rows": len(products),
        "unique_rows": len(deduped),
        "duplicate_rows": duplicate_count,
        "duplicate_samples": duplicate_samples,
    }


def validate_rest_payload_columns(column_names, trace=None):
    import httpx

    accepted_columns = list(column_names)
    skipped_columns = []

    while accepted_columns:
        response = httpx.get(
            supabase_table_url(),
            headers=supabase_headers(),
            params={
                "select": ",".join(accepted_columns),
                "limit": "0",
            },
            timeout=SUPABASE_TIMEOUT_SECONDS,
        )

        if response.status_code < 400:
            return accepted_columns, skipped_columns

        missing_column = missing_column_from_postgrest_response(response)
        if missing_column and missing_column in accepted_columns:
            accepted_columns.remove(missing_column)
            skipped_columns.append(missing_column)
            trace_step(trace, "skip_missing_supabase_column", column=missing_column)
            continue

        raise RuntimeError(f"Supabase schema check failed: {response_preview(response)}")

    return accepted_columns, skipped_columns


def write_products_to_supabase_rest(products, trace=None):
    if not supabase_rest_configured():
        return {
            "enabled": False,
            "reason": "SUPABASE_URL_or_SUPABASE_SERVICE_ROLE_KEY_missing",
            "scraped_rows": len(products),
        }

    import httpx

    schema_name = validate_identifier(products_schema(), "CANNAFLOW_PRODUCTS_SCHEMA")
    table_name = validate_identifier(products_table(), "CANNAFLOW_PRODUCTS_TABLE")
    upsert_on = products_upsert_on()
    replace_all = products_replace_all()

    if upsert_on:
        validate_identifier(upsert_on, "CANNAFLOW_PRODUCTS_UPSERT_ON")

    trace_step(
        trace,
        "write_products_to_supabase_rest",
        schema=schema_name,
        table=table_name,
        rows=len(products),
        replace_all=replace_all,
        upsert_on=upsert_on,
    )

    payloads = [{key: json_ready(value) for key, value in product.items()} for product in products]
    insert_columns = []
    skipped_columns = []

    if payloads:
        insert_columns, skipped_columns = validate_rest_payload_columns(list(payloads[0].keys()), trace=trace)
        payloads = [{column: payload.get(column) for column in insert_columns} for payload in payloads]

    deleted_rows = None

    if replace_all:
        response = httpx.delete(
            supabase_table_url(),
            headers={
                **supabase_headers(),
                "Prefer": "return=representation,count=exact",
            },
            params={"product_name": "not.is.null"},
            timeout=SUPABASE_TIMEOUT_SECONDS,
        )

        if response.status_code >= 400:
            raise RuntimeError(f"Supabase cleanup failed: {response_preview(response)}")

        try:
            deleted_payload = response.json()
            if isinstance(deleted_payload, list):
                deleted_rows = len(deleted_payload)
        except Exception:
            deleted_rows = None

    insert_url = supabase_table_url()
    prefer_header = "return=minimal"

    if upsert_on:
        insert_url = f"{insert_url}?on_conflict={quote(upsert_on, safe=',')}"
        prefer_header = "resolution=merge-duplicates,return=minimal"

    inserted_rows = 0
    for product_chunk in chunks(payloads, 500):
        if not product_chunk:
            continue

        response = httpx.post(
            insert_url,
            headers={
                **supabase_headers(),
                "Prefer": prefer_header,
            },
            json=product_chunk,
            timeout=SUPABASE_TIMEOUT_SECONDS,
        )

        if response.status_code >= 400:
            raise RuntimeError(f"Supabase insert failed: {response_preview(response)}")

        inserted_rows += len(product_chunk)

    return {
        "enabled": True,
        "method": "supabase_rest",
        "schema": schema_name,
        "table": table_name,
        "replace_all": replace_all,
        "deleted_rows": deleted_rows,
        "upsert_on": upsert_on,
        "inserted_rows": inserted_rows,
        "insert_columns": insert_columns,
        "skipped_missing_columns": skipped_columns,
        "supabase_url": supabase_url(),
    }


def fetch_previous_products_snapshot(trace=None):
    if not stock_report_enabled() or not supabase_rest_configured():
        return []

    import httpx

    select_columns = ",".join(
        [
            "product_name",
            "available_grams",
            "available_text",
            "stock_grams",
            "stock_text",
            "sale_price_text",
            "sale_price_per_g",
            "product_type",
            "cultivar",
            "manufacturer",
            "sale_enabled",
            "expiry_date_text",
            "expiry_date",
            "scraped_at",
        ]
    )
    trace_step(trace, "fetch_previous_products_supabase_rest", table=products_table())
    response = httpx.get(
        supabase_table_url(),
        headers=supabase_headers(),
        params={
            "select": select_columns,
            "product_name": "not.is.null",
            "limit": "5000",
        },
        timeout=SUPABASE_TIMEOUT_SECONDS,
    )

    if response.status_code >= 400:
        raise RuntimeError(f"Supabase previous snapshot fetch failed: {response_preview(response)}")

    payload = response.json()
    if not isinstance(payload, list):
        raise RuntimeError("Supabase previous snapshot response was not a list.")

    return payload


def compare_stock_snapshots(previous_products, current_products):
    previous_by_name = {
        product_name_key(product): product
        for product in previous_products
        if product_name_key(product)
    }
    current_by_name = {
        product_name_key(product): product
        for product in current_products
        if product_name_key(product)
    }
    changes = []

    for key in sorted(current_by_name, key=lambda item: normalize_space(current_by_name[item].get("product_name")).casefold()):
        current = current_by_name[key]
        previous = previous_by_name.get(key)
        current_available = decimal_for_compare(current.get("available_grams"))
        previous_available = decimal_for_compare(previous.get("available_grams")) if previous else None
        current_text = quantity_display(current.get("available_grams"), current.get("available_text"))
        previous_text = quantity_display(
            previous.get("available_grams") if previous else None,
            previous.get("available_text") if previous else None,
        )
        change_type = "new" if previous is None else "changed"

        if previous is not None:
            if current_available is not None and previous_available is not None:
                if current_available == previous_available:
                    continue
            elif current_text == previous_text:
                continue

        difference = None
        if current_available is not None and previous_available is not None:
            difference = current_available - previous_available

        changes.append(
            {
                "change_type": change_type,
                "product_name": current.get("product_name"),
                "previous_available": previous_available,
                "current_available": current_available,
                "previous_available_text": previous_text,
                "current_available_text": current_text,
                "difference": difference,
                "previous_stock": decimal_for_compare(previous.get("stock_grams")) if previous else None,
                "current_stock": decimal_for_compare(current.get("stock_grams")),
                "previous_stock_text": quantity_display(
                    previous.get("stock_grams") if previous else None,
                    previous.get("stock_text") if previous else None,
                ),
                "current_stock_text": quantity_display(current.get("stock_grams"), current.get("stock_text")),
                "manufacturer": current.get("manufacturer"),
                "cultivar": current.get("cultivar"),
                "product_type": current.get("product_type"),
                "sale_price_text": current.get("sale_price_text"),
                "sale_enabled": current.get("sale_enabled"),
                "expiry_date_text": current.get("expiry_date_text"),
                "page_number": current.get("page_number"),
                "row_index": current.get("row_index"),
                "previous_scraped_at": previous.get("scraped_at") if previous else None,
                "current_scraped_at": current.get("scraped_at"),
            }
        )

    for key in sorted(set(previous_by_name) - set(current_by_name), key=lambda item: normalize_space(previous_by_name[item].get("product_name")).casefold()):
        previous = previous_by_name[key]
        changes.append(
            {
                "change_type": "removed",
                "product_name": previous.get("product_name"),
                "previous_available": decimal_for_compare(previous.get("available_grams")),
                "current_available": None,
                "previous_available_text": quantity_display(previous.get("available_grams"), previous.get("available_text")),
                "current_available_text": "",
                "difference": None,
                "previous_stock": decimal_for_compare(previous.get("stock_grams")),
                "current_stock": None,
                "previous_stock_text": quantity_display(previous.get("stock_grams"), previous.get("stock_text")),
                "current_stock_text": "",
                "manufacturer": previous.get("manufacturer"),
                "cultivar": previous.get("cultivar"),
                "product_type": previous.get("product_type"),
                "sale_price_text": previous.get("sale_price_text"),
                "sale_enabled": previous.get("sale_enabled"),
                "expiry_date_text": previous.get("expiry_date_text"),
                "page_number": None,
                "row_index": None,
                "previous_scraped_at": previous.get("scraped_at"),
                "current_scraped_at": None,
            }
        )

    decreased = [
        change
        for change in changes
        if isinstance(change.get("difference"), Decimal) and change["difference"] < 0
    ]
    increased = [
        change
        for change in changes
        if isinstance(change.get("difference"), Decimal) and change["difference"] > 0
    ]

    return changes, {
        "previous_rows": len(previous_products),
        "current_rows": len(current_products),
        "changed_rows": len(changes),
        "decreased_rows": len(decreased),
        "increased_rows": len(increased),
        "new_rows": sum(1 for change in changes if change.get("change_type") == "new"),
        "removed_rows": sum(1 for change in changes if change.get("change_type") == "removed"),
    }


def auto_width_for_sheet(sheet, max_width=60):
    for column_cells in sheet.columns:
        column_letter = column_cells[0].column_letter
        max_length = 0

        for cell in column_cells:
            value = cell.value
            if value is None:
                continue
            max_length = max(max_length, len(str(value)))

        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), max_width)


def create_stock_change_excel_report(changes, current_products, previous_count, scraped_at, timestamp, trace=None):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    os.makedirs(ARTIFACTS_DIR, exist_ok=True)
    report_path = os.path.join(ARTIFACTS_DIR, f"cannaflow-stock-changes-{timestamp}.xlsx")
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Zusammenfassung"
    changes_sheet = workbook.create_sheet("Mengenänderungen")
    current_sheet = workbook.create_sheet("Aktueller Bestand")

    changed_rows = len(changes)
    decreased_rows = sum(
        1
        for change in changes
        if isinstance(change.get("difference"), Decimal) and change["difference"] < 0
    )
    increased_rows = sum(
        1
        for change in changes
        if isinstance(change.get("difference"), Decimal) and change["difference"] > 0
    )
    new_rows = sum(1 for change in changes if change.get("change_type") == "new")
    removed_rows = sum(1 for change in changes if change.get("change_type") == "removed")

    summary_rows = [
        ["Cannaflow Mengenänderungen", ""],
        ["Erstellt am", str(scraped_at)],
        ["Vorherige Produkte", previous_count],
        ["Aktuelle Produkte", len(current_products)],
        ["Zeilen im Report", changed_rows],
        ["Verfügbar gesunken", decreased_rows],
        ["Verfügbar gestiegen", increased_rows],
        ["Neue Produkte", new_rows],
        ["Nicht mehr im aktuellen Inventar", removed_rows],
        ["Verglichene Spalte", "Verfügbar"],
    ]
    summary_sheet.append(summary_rows[0])
    summary_sheet["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    summary_sheet["A1"].fill = PatternFill("solid", fgColor="1F4E5F")
    summary_sheet["B1"].fill = PatternFill("solid", fgColor="1F4E5F")
    for row in summary_rows[1:]:
        summary_sheet.append(row)
    summary_sheet.column_dimensions["A"].width = 34
    summary_sheet.column_dimensions["B"].width = 34

    headers = [
        "Produkt",
        "Typ",
        "Verfügbar alt",
        "Verfügbar neu",
        "Änderung (neu - alt)",
        "Bestand alt",
        "Bestand neu",
        "Hersteller",
        "Kultivar",
        "Art",
        "Preis pro g",
        "Verkauf",
        "Verfall",
        "Seite",
        "Zeile",
        "Vorheriger Lauf",
        "Aktueller Lauf",
    ]
    changes_sheet.append(headers)
    if changes:
        for change in changes:
            changes_sheet.append(
                [
                    change.get("product_name"),
                    change.get("change_type"),
                    excel_number(change.get("previous_available"))
                    if change.get("previous_available") is not None
                    else change.get("previous_available_text"),
                    excel_number(change.get("current_available"))
                    if change.get("current_available") is not None
                    else change.get("current_available_text"),
                    excel_number(change.get("difference")) if change.get("difference") is not None else "",
                    excel_number(change.get("previous_stock"))
                    if change.get("previous_stock") is not None
                    else change.get("previous_stock_text"),
                    excel_number(change.get("current_stock"))
                    if change.get("current_stock") is not None
                    else change.get("current_stock_text"),
                    change.get("manufacturer"),
                    change.get("cultivar"),
                    change.get("product_type"),
                    change.get("sale_price_text"),
                    change.get("sale_enabled"),
                    change.get("expiry_date_text"),
                    change.get("page_number"),
                    change.get("row_index"),
                    str(change.get("previous_scraped_at") or ""),
                    str(change.get("current_scraped_at") or ""),
                ]
            )
    else:
        changes_sheet.append(["Keine Mengenänderungen seit dem letzten Lauf.", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""])

    current_headers = [
        "Produkt",
        "Verfügbar",
        "Bestand",
        "Hersteller",
        "Kultivar",
        "Art",
        "Preis pro g",
        "Verkauf",
        "Verfall",
        "Seite",
        "Zeile",
        "Lauf",
    ]
    current_sheet.append(current_headers)
    for product in sorted(current_products, key=lambda item: normalize_space(item.get("product_name")).casefold()):
        current_sheet.append(
            [
                product.get("product_name"),
                excel_number(product.get("available_grams")) if product.get("available_grams") is not None else product.get("available_text"),
                excel_number(product.get("stock_grams")) if product.get("stock_grams") is not None else product.get("stock_text"),
                product.get("manufacturer"),
                product.get("cultivar"),
                product.get("product_type"),
                product.get("sale_price_text"),
                product.get("sale_enabled"),
                product.get("expiry_date_text"),
                product.get("page_number"),
                product.get("row_index"),
                str(product.get("scraped_at") or ""),
            ]
        )

    header_fill = PatternFill("solid", fgColor="1F4E5F")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(bottom=Side(style="thin", color="B7B7B7"))
    decrease_fill = PatternFill("solid", fgColor="FCE4D6")
    increase_fill = PatternFill("solid", fgColor="E2F0D9")
    new_fill = PatternFill("solid", fgColor="DDEBF7")
    removed_fill = PatternFill("solid", fgColor="E7E6E6")

    for sheet in [changes_sheet, current_sheet]:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.border = thin_border
            if sheet.title == "Mengenänderungen":
                difference = row[4].value
                change_type = row[1].value
                fill = None
                if isinstance(difference, (int, float)) and difference < 0:
                    fill = decrease_fill
                elif isinstance(difference, (int, float)) and difference > 0:
                    fill = increase_fill
                elif change_type == "new":
                    fill = new_fill
                elif change_type == "removed":
                    fill = removed_fill
                if fill:
                    for cell in row:
                        cell.fill = fill
        auto_width_for_sheet(sheet)

    for sheet in [changes_sheet, current_sheet]:
        numeric_columns = [3, 4, 5, 6, 7] if sheet.title == "Mengenänderungen" else [2, 3]
        for column_index in numeric_columns:
            column_letter = get_column_letter(column_index)
            for cell in sheet[column_letter][1:]:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '0.##'

    workbook.save(report_path)
    trace_step(
        trace,
        "created_stock_change_excel_report",
        path=report_path,
        changed_rows=changed_rows,
        current_rows=len(current_products),
        previous_rows=previous_count,
    )

    return report_path


def build_stock_report(previous_products, current_products, scraped_at, timestamp, base_url, trace=None):
    if not stock_report_enabled():
        return {
            "enabled": False,
            "reason": "CANNAFLOW_STOCK_REPORT_ENABLED=false",
        }

    changes, summary = compare_stock_snapshots(previous_products, current_products)
    report_path = create_stock_change_excel_report(
        changes,
        current_products,
        len(previous_products),
        scraped_at,
        timestamp,
        trace=trace,
    )

    return {
        "enabled": True,
        **summary,
        "filename": os.path.basename(report_path),
        "path": report_path,
        **report_response_links(base_url, report_path),
    }


def scrape_inventory_products(base_url, trace=None):
    os.makedirs(ARTIFACTS_DIR, exist_ok=True)
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    screenshot_path = os.path.join(ARTIFACTS_DIR, f"cannaflow-scrape-products-{timestamp}.png")
    scraped_at = utc_now_iso()
    raw_rows = []
    pages_scraped = []

    trace_step(trace, "start_browser", headless=bool_env("CANNAFLOW_HEADLESS", True))
    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=bool_env("CANNAFLOW_HEADLESS", True))
        context = None
        try:
            context, page, reused_session = open_authenticated_inventory(browser, trace=trace)
            page_size_result = ensure_inventory_page_size(page, trace=trace)

            for page_index in range(1, SCRAPE_MAX_PAGES + 1):
                pagination_before = get_pagination_state(page)
                current_page = page_index
                scrape_result = scrape_current_inventory_page(page, page_number=current_page, trace=trace)
                page_rows = scrape_result.get("rows", [])

                for row in page_rows:
                    row["page_number"] = current_page
                    row["source_url"] = page.url
                    raw_rows.append(row)

                pages_scraped.append(
                    {
                        "page": current_page,
                        "label": pagination_before.get("label"),
                        "rows": len(page_rows),
                        "range_from": pagination_before.get("from_row"),
                        "range_to": pagination_before.get("to_row"),
                        "total_rows": pagination_before.get("total_rows"),
                    }
                )

                if not pagination_before.get("has_next"):
                    break

                click_result = click_next_inventory_page(page, pagination_before, trace=trace)
                if not click_result.get("clicked"):
                    break
            else:
                raise RuntimeError(f"Stopped after CANNAFLOW_SCRAPE_MAX_PAGES={SCRAPE_MAX_PAGES}")

            normalized_products = [
                normalize_scraped_product(raw_row, raw_row.get("page_number"), raw_row.get("source_url") or page.url, scraped_at)
                for raw_row in raw_rows
                if normalize_space(raw_row.get("product_name"))
            ]
            products, dedupe = dedupe_products_by_name(normalized_products)

            if raw_rows and not products:
                raise RuntimeError("Scrape found table rows, but no product names were extracted.")

            previous_products = fetch_previous_products_snapshot(trace=trace)
            database_result = write_products_to_supabase_rest(products, trace=trace)
            stock_report_result = build_stock_report(
                previous_products,
                products,
                scraped_at,
                timestamp,
                base_url,
                trace=trace,
            )

            trace_step(trace, "capture_scrape_screenshot", path=screenshot_path)
            page.screenshot(path=screenshot_path, full_page=True)

            return {
                "ok": True,
                "current_url": page.url,
                "page_title": page.title(),
                "reused_session": reused_session,
                "ready_text": ready_text(),
                "page_size": page_size_result,
                "pages_scraped": pages_scraped,
                "raw_rows_seen": len(raw_rows),
                "products_scraped": len(products),
                "dedupe": dedupe,
                "database": database_result,
                "stock_report": stock_report_result,
                "screenshot_path": screenshot_path,
                **screenshot_response_links(base_url, screenshot_path),
            }
        finally:
            if context:
                context.close()
            browser.close()


def login_only(base_url, trace=None):
    os.makedirs(ARTIFACTS_DIR, exist_ok=True)
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    screenshot_path = os.path.join(ARTIFACTS_DIR, f"cannaflow-after-login-{timestamp}.png")
    wait_ms = int_env("CANNAFLOW_AFTER_LOGIN_WAIT_MS", 5_000)

    trace_step(trace, "start_browser", headless=bool_env("CANNAFLOW_HEADLESS", True))
    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=bool_env("CANNAFLOW_HEADLESS", True))
        context = None
        try:
            context, page, reused_session = open_authenticated_inventory(browser, trace=trace)
            page_size_result = ensure_inventory_page_size(page, trace=trace)
            trace_step(trace, "wait_before_screenshot", wait_ms=wait_ms)
            page.wait_for_timeout(wait_ms)
            trace_step(trace, "capture_screenshot", path=screenshot_path)
            page.screenshot(path=screenshot_path, full_page=True)

            return {
                "ok": True,
                "current_url": page.url,
                "page_title": page.title(),
                "reused_session": reused_session,
                "ready_text": ready_text(),
                "page_size": page_size_result,
                "session_state_path": SESSION_STATE_PATH,
                "session_state_exists": os.path.exists(SESSION_STATE_PATH),
                "screenshot_path": screenshot_path,
                "screenshot_wait_ms": wait_ms,
                **screenshot_response_links(base_url, screenshot_path),
            }
        finally:
            if context:
                context.close()
            browser.close()


@app.get("/health")
def health():
    return {
        "ok": True,
        "service": SERVICE_NAME,
        "uptime_seconds": round(time.time() - STARTED_AT, 3),
        "login_url_configured": bool((os.environ.get("CANNAFLOW_LOGIN_URL") or "").strip()),
        "inventory_url_configured": bool((os.environ.get("CANNAFLOW_INVENTORY_URL") or "").strip()),
        "session_state_path": SESSION_STATE_PATH,
        "session_state_exists": os.path.exists(SESSION_STATE_PATH),
        "ready_text": ready_text(),
        "page_size_target": inventory_page_size(),
        "screenshot_wait_ms": int_env("CANNAFLOW_AFTER_LOGIN_WAIT_MS", 5_000),
        "ready_timeout_ms": int_env("CANNAFLOW_READY_TIMEOUT_MS", 60_000),
        "scrape_max_pages": SCRAPE_MAX_PAGES,
        "supabase_rest_configured": supabase_rest_configured(),
        "products_schema": products_schema(),
        "products_table": products_table(),
        "products_replace_all": products_replace_all(),
        "products_upsert_on": products_upsert_on(),
        "stock_report_enabled": stock_report_enabled(),
    }


@app.post("/jobs/login")
def login_job(request: Request):
    base_url = base_url_from_request(request)
    try:
        return login_only(base_url)
    except Exception as exc:
        details = getattr(exc, "details", {}) or {}
        debug_screenshot_path = details.get("debug_screenshot_path")
        if debug_screenshot_path:
            details.update(screenshot_response_links(base_url, debug_screenshot_path))

        return JSONResponse(
            status_code=500,
            content={
                "ok": False,
                "error": f"{type(exc).__name__}: {exc}",
                **details,
                "traceback": traceback.format_exc().splitlines()[-12:],
            },
        )


@app.post("/jobs/login/run")
def login_job_run(request: Request):
    return login_job(request)


@app.post("/jobs/scrape-products")
def scrape_products_job(request: Request):
    base_url = base_url_from_request(request)
    try:
        return scrape_inventory_products(base_url)
    except Exception as exc:
        details = getattr(exc, "details", {}) or {}
        debug_screenshot_path = details.get("debug_screenshot_path")
        if debug_screenshot_path:
            details.update(screenshot_response_links(base_url, debug_screenshot_path))

        return JSONResponse(
            status_code=500,
            content={
                "ok": False,
                "error": f"{type(exc).__name__}: {exc}",
                **details,
                "traceback": traceback.format_exc().splitlines()[-12:],
            },
        )


@app.post("/jobs/scrape-products/run")
def scrape_products_job_run(request: Request):
    return scrape_products_job(request)


@app.get("/screenshots/latest", name="get_latest_screenshot")
def get_latest_screenshot():
    screenshot_path = latest_screenshot_path()
    if not screenshot_path:
        return JSONResponse(status_code=404, content={"ok": False, "error": "no_screenshot_found"})
    return FileResponse(screenshot_path, media_type="image/png", filename=os.path.basename(screenshot_path))


@app.get("/screenshots/{filename}", name="get_screenshot")
def get_screenshot(filename: str):
    screenshot_path = os.path.join(ARTIFACTS_DIR, safe_screenshot_filename(filename))
    if not os.path.exists(screenshot_path):
        return JSONResponse(status_code=404, content={"ok": False, "error": "screenshot_not_found"})
    return FileResponse(screenshot_path, media_type="image/png", filename=os.path.basename(screenshot_path))


@app.get("/reports/latest", name="get_latest_report")
def get_latest_report():
    report_path = latest_report_path()
    if not report_path:
        return JSONResponse(status_code=404, content={"ok": False, "error": "no_report_found"})
    return FileResponse(
        report_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=os.path.basename(report_path),
    )


@app.get("/reports/{filename}", name="get_report")
def get_report(filename: str):
    try:
        safe_filename = safe_report_filename(filename)
    except RuntimeError as exc:
        return JSONResponse(status_code=400, content={"ok": False, "error": str(exc)})

    report_path = os.path.join(ARTIFACTS_DIR, safe_filename)
    if not os.path.exists(report_path):
        return JSONResponse(status_code=404, content={"ok": False, "error": "report_not_found"})
    return FileResponse(
        report_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=safe_filename,
    )
