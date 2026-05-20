import os
import json
import mimetypes
import re
import threading
import time
import traceback
import uuid
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from urllib.parse import quote
from zoneinfo import ZoneInfo

from fastapi import FastAPI, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse
from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
from playwright.sync_api import sync_playwright


STARTED_AT = time.time()
SERVICE_NAME = "wawican-scan"
ARTIFACTS_DIR = os.environ.get("ARTIFACTS_DIR", "/app/artifacts")
SESSION_STATE_PATH = os.environ.get(
    "WAWICAN_SESSION_STATE_PATH",
    os.path.join(ARTIFACTS_DIR, "wawican-storage-state.json"),
)
DEFAULT_USER_AGENT = (
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/136.0.0.0 Safari/537.36"
)
WAWICAN_USER_AGENT = (
    os.environ.get("WAWICAN_USER_AGENT", DEFAULT_USER_AGENT).strip()
    or DEFAULT_USER_AGENT
)
INVENTORY_READY_TIMEOUT_MS = int(os.environ.get("WAWICAN_INVENTORY_READY_TIMEOUT_MS", "60000"))
POST_LOGIN_READY_TIMEOUT_MS = int(os.environ.get("WAWICAN_POST_LOGIN_READY_TIMEOUT_MS", "15000"))
FILTER_TIMEOUT_MS = int(os.environ.get("WAWICAN_FILTER_TIMEOUT_MS", "60000"))
SCRAPE_PAGE_READY_TIMEOUT_MS = int(os.environ.get("WAWICAN_SCRAPE_PAGE_READY_TIMEOUT_MS", "30000"))
SCRAPE_MAX_PAGES = int(os.environ.get("WAWICAN_SCRAPE_MAX_PAGES", "250"))
SUPABASE_TIMEOUT_SECONDS = int(os.environ.get("WAWICAN_SUPABASE_TIMEOUT_SECONDS", "60"))
MAX_STORED_JOBS = 50
JOB_LOCK = threading.Lock()
JOBS = {}
GERMAN_TIMEZONE = ZoneInfo(os.environ.get("TZ", "Europe/Berlin") or "Europe/Berlin")


app = FastAPI(title=SERVICE_NAME)

cors_origins = [
    origin.strip()
    for origin in os.environ.get(
        "WAWICAN_SCAN_CORS_ORIGINS",
        "http://localhost:8040,http://127.0.0.1:8040,http://178.104.144.30:8040",
    ).split(",")
    if origin.strip()
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=cors_origins,
    allow_methods=["GET", "POST", "OPTIONS"],
    allow_headers=["*"],
)


def bool_env(name, default=True):
    value = os.environ.get(name)

    if value is None:
        return default

    return value.strip().lower() in {"1", "true", "yes", "on"}


def int_env(name, default):
    value = os.environ.get(name)

    if value is None or not value.strip():
        return default

    try:
        return int(value)
    except ValueError as exc:
        raise RuntimeError(f"Invalid integer environment variable {name}={value!r}") from exc


def required_env(name):
    value = (os.environ.get(name) or "").strip()

    if not value:
        raise RuntimeError(f"Missing required environment variable: {name}")

    return value


def inventory_url():
    return (os.environ.get("WAWICAN_INVENTORY_URL") or "").strip() or required_env("WAWICAN_LOGIN_URL")


def login_url():
    return (os.environ.get("WAWICAN_LOGIN_URL") or "").strip() or inventory_url()


def ready_text():
    return (os.environ.get("WAWICAN_READY_TEXT") or "Verfügbarkeit").strip() or "Verfügbarkeit"


def browser_context_options():
    return {
        "user_agent": WAWICAN_USER_AGENT,
        "viewport": {"width": 1365, "height": 900},
        "locale": "de-DE",
        "timezone_id": os.environ.get("TZ", "Europe/Berlin"),
    }


def screenshot_wait_ms():
    return 0


def safe_artifact_filename(filename):
    if not re.match(r"^[A-Za-z0-9_.-]+\.png$", filename or ""):
        raise RuntimeError("Invalid screenshot filename.")

    return filename


def safe_report_filename(filename):
    if not re.match(r"^[A-Za-z0-9_.-]+\.xlsx$", filename or ""):
        raise RuntimeError("Invalid report filename.")

    return filename


def latest_screenshot_path():
    if not os.path.isdir(ARTIFACTS_DIR):
        return None

    screenshot_paths = [
        os.path.join(ARTIFACTS_DIR, filename)
        for filename in os.listdir(ARTIFACTS_DIR)
        if filename.lower().endswith(".png")
    ]

    if not screenshot_paths:
        return None

    return max(screenshot_paths, key=os.path.getmtime)


def latest_report_path():
    if not os.path.isdir(ARTIFACTS_DIR):
        return None

    report_paths = [
        os.path.join(ARTIFACTS_DIR, filename)
        for filename in os.listdir(ARTIFACTS_DIR)
        if filename.lower().endswith(".xlsx")
    ]

    if not report_paths:
        return None

    return max(report_paths, key=os.path.getmtime)


def utc_now_iso():
    return datetime.now(timezone.utc).isoformat()


def base_url_from_request(request):
    return str(request.base_url).rstrip("/")


def screenshot_file_url(base_url, screenshot_path):
    filename = os.path.basename(screenshot_path)
    return f"{base_url}/screenshots/{filename}"


def screenshot_response_links(base_url, screenshot_path):
    return {
        "screenshot_url": screenshot_file_url(base_url, screenshot_path),
        "latest_screenshot_url": f"{base_url}/screenshots/latest",
    }


def report_file_url(base_url, report_path):
    filename = os.path.basename(report_path)
    return f"{base_url}/reports/{filename}"


def report_response_links(base_url, report_path):
    return {
        "report_url": report_file_url(base_url, report_path),
        "latest_report_url": f"{base_url}/reports/latest",
    }


def trace_step(trace, name, **fields):
    if trace:
        trace(name, **fields)


def capture_screenshot_now(page, path, trace=None, trigger="inventory_ready_visible"):
    trace_step(trace, "capture_screenshot", path=path)
    page.screenshot(path=path, full_page=True)

    return {
        "screenshot_path": path,
        "screenshot_wait_ms": 0,
        "screenshot_trigger": trigger,
    }


def page_text_excerpt(page, limit=500):
    try:
        text = page.locator("body").inner_text(timeout=2_000)
    except Exception:
        return ""

    text = re.sub(r"\s+", " ", text or "").strip()
    return text[:limit]


def trace_page_state(trace, name, page, **fields):
    try:
        title = page.title()
    except Exception:
        title = ""

    trace_step(
        trace,
        name,
        current_url=page.url,
        page_title=title,
        body_excerpt=page_text_excerpt(page),
        **fields,
    )


def capture_debug_screenshot(page, path, trace, step_name):
    try:
        page.screenshot(path=path, full_page=True)
        trace_step(trace, step_name, path=path, current_url=page.url)
    except Exception as exc:
        trace_step(trace, f"{step_name}_failed", error=f"{type(exc).__name__}: {exc}")


def first_visible_locator(page, selectors, timeout=5_000):
    for selector in selectors:
        selector = (selector or "").strip()
        if not selector:
            continue

        locator = page.locator(selector).first
        try:
            locator.wait_for(state="visible", timeout=timeout)
            return locator
        except PlaywrightTimeoutError:
            continue

    return None


def wait_for_next_render_frame(page):
    page.evaluate(
        """
        () => new Promise((resolve) => {
          requestAnimationFrame(() => requestAnimationFrame(resolve));
        })
        """
    )


def wait_for_inventory_page(page, timeout=INVENTORY_READY_TIMEOUT_MS, trace=None):
    trace_step(trace, "wait_for_inventory_page", timeout_ms=timeout, ready_text=ready_text())
    page.wait_for_function(
        """
        (targetText) => {
          const fold = (value) => (value || '')
            .normalize('NFD')
            .replace(/[\\u0300-\\u036f]/g, '')
            .replace(/\\s+/g, ' ')
            .trim()
            .toLowerCase();
          const target = fold(targetText);
          const selectors = [
            'th',
            '[role="columnheader"]',
            'span',
            'div',
            'button',
            'label',
            '.q-table *',
            '.inventory-table-component *'
          ];
          const candidates = Array.from(document.querySelectorAll(selectors.join(',')));
          return candidates.some((element) => {
            const text = fold(element.textContent);
            if (!text.includes(target)) {
              return false;
            }

            const rect = element.getBoundingClientRect();
            const style = window.getComputedStyle(element);

            return (
              rect.width > 0 &&
              rect.height > 0 &&
              style.visibility !== 'hidden' &&
              style.display !== 'none' &&
              style.opacity !== '0'
            );
          });
        }
        """,
        arg=ready_text(),
        timeout=timeout,
    )
    wait_for_next_render_frame(page)
    trace_page_state(trace, "inventory_ready_visible", page)

    return True


def login_form_is_visible(page):
    username_selector = (os.environ.get("WAWICAN_USERNAME_SELECTOR") or "").strip()
    password_selector = (os.environ.get("WAWICAN_PASSWORD_SELECTOR") or "").strip()

    username = first_visible_locator(
        page,
        [
            username_selector,
            '[data-testid="login-email-input"]',
            'input[aria-label*="E-Mail" i]',
            'input[aria-label*="Mail" i]',
            'input[placeholder*="Email" i]',
            'input[placeholder*="E-Mail" i]',
            'input[placeholder*="Benutzer" i]',
            'input[placeholder*="Username" i]',
            'input[type="email"]',
            'input[name*="email" i]',
            'input[name*="user" i]',
        ],
        timeout=1_000,
    )
    password = first_visible_locator(
        page,
        [
            password_selector,
            '[data-testid="login-password-input"]',
            'input[aria-label*="Passwort" i]',
            'input[placeholder*="Password" i]',
            'input[placeholder*="Passwort" i]',
            'input[type="password"]',
            'input[name*="password" i]',
            'input[name*="passwort" i]',
        ],
        timeout=1_000,
    )

    return username is not None and password is not None


def fill_login_form(page):
    username = required_env("WAWICAN_USERNAME")
    password = required_env("WAWICAN_PASSWORD")
    username_selector = (os.environ.get("WAWICAN_USERNAME_SELECTOR") or "").strip()
    password_selector = (os.environ.get("WAWICAN_PASSWORD_SELECTOR") or "").strip()

    username_input = first_visible_locator(
        page,
        [
            username_selector,
            '[data-testid="login-email-input"]',
            'input[aria-label*="E-Mail" i]',
            'input[aria-label*="Mail" i]',
            'input[placeholder*="Email" i]',
            'input[placeholder*="E-Mail" i]',
            'input[placeholder*="Benutzer" i]',
            'input[placeholder*="Username" i]',
            'input[type="email"]',
            'input[name*="email" i]',
            'input[name*="user" i]',
            'input:not([type="hidden"])',
        ],
    )
    password_input = first_visible_locator(
        page,
        [
            password_selector,
            '[data-testid="login-password-input"]',
            'input[aria-label*="Passwort" i]',
            'input[placeholder*="Password" i]',
            'input[placeholder*="Passwort" i]',
            'input[type="password"]',
            'input[name*="password" i]',
            'input[name*="passwort" i]',
        ],
    )

    if username_input is None or password_input is None:
        raise RuntimeError(
            "Could not find login inputs. Set WAWICAN_USERNAME_SELECTOR and "
            "WAWICAN_PASSWORD_SELECTOR in Dokploy."
        )

    username_input.fill(username)
    password_input.fill(password)

    extra_click_text = (os.environ.get("WAWICAN_LOGIN_EXTRA_CLICK_TEXT") or "").strip()
    if extra_click_text:
        try:
            page.get_by_text(extra_click_text, exact=True).click(timeout=5_000)
        except PlaywrightTimeoutError:
            print(f"extra login click text was not visible: {extra_click_text!r}", flush=True)


def submit_login_form(page):
    login_button_selector = (os.environ.get("WAWICAN_LOGIN_BUTTON_SELECTOR") or "").strip()

    button = first_visible_locator(
        page,
        [
            login_button_selector,
            'button[type="submit"]',
            'button:has-text("Login")',
            'button:has-text("Log in")',
            'button:has-text("Einloggen")',
            'button:has-text("Anmelden")',
            'input[type="submit"]',
        ],
        timeout=5_000,
    )

    if button is not None:
        button.click(timeout=10_000)
    else:
        page.keyboard.press("Enter")


def wait_for_inventory_after_login_submit(page, trace=None):
    try:
        trace_step(trace, "wait_for_inventory_after_submit", timeout_ms=POST_LOGIN_READY_TIMEOUT_MS)
        wait_for_inventory_page(page, timeout=POST_LOGIN_READY_TIMEOUT_MS, trace=trace)
        return
    except Exception as exc:
        trace_page_state(
            trace,
            "inventory_not_visible_after_submit",
            page,
            error=f"{type(exc).__name__}: {exc}",
        )

    trace_step(trace, "goto_inventory_after_login", url=inventory_url())
    page.goto(inventory_url(), wait_until="domcontentloaded", timeout=30_000)
    wait_for_inventory_page(page, trace=trace)


def save_session_state(context):
    session_state_dir = os.path.dirname(SESSION_STATE_PATH)

    if session_state_dir:
        os.makedirs(session_state_dir, exist_ok=True)

    temp_path = f"{SESSION_STATE_PATH}.{os.getpid()}.tmp"
    try:
        context.storage_state(path=temp_path)
        os.replace(temp_path, SESSION_STATE_PATH)
    finally:
        if os.path.exists(temp_path):
            os.remove(temp_path)


def open_saved_session(browser, trace=None):
    if not os.path.exists(SESSION_STATE_PATH):
        trace_step(trace, "saved_session_missing", path=SESSION_STATE_PATH)
        return None

    trace_step(trace, "open_saved_session", path=SESSION_STATE_PATH)
    print("trying saved Wawican session ...", flush=True)

    try:
        context = browser.new_context(
            storage_state=SESSION_STATE_PATH,
            **browser_context_options(),
        )
    except Exception as exc:
        print(f"saved Wawican session could not be opened: {type(exc).__name__}: {exc}", flush=True)
        return None

    page = context.new_page()

    try:
        trace_step(trace, "goto_inventory_with_saved_session", url=inventory_url())
        page.goto(inventory_url(), wait_until="domcontentloaded", timeout=30_000)
        page.wait_for_load_state("domcontentloaded", timeout=10_000)
        wait_for_inventory_page(page, trace=trace)
        trace_step(trace, "saved_session_valid")
        return context, page, True, True
    except Exception as exc:
        context.close()
        trace_step(trace, "saved_session_invalid", error=f"{type(exc).__name__}: {exc}")
        print(f"saved Wawican session is expired or invalid: {type(exc).__name__}: {exc}", flush=True)
        return None


def open_fresh_session(browser, before_login_path=None, trace=None):
    trace_step(trace, "open_fresh_session")
    print("trying fresh Wawican login ...", flush=True)

    context = browser.new_context(**browser_context_options())
    page = context.new_page()

    trace_step(trace, "goto_login_url", url=login_url())
    page.goto(login_url(), wait_until="domcontentloaded", timeout=30_000)
    page.wait_for_load_state("domcontentloaded", timeout=10_000)
    trace_page_state(trace, "login_page_loaded", page)

    login_form_visible = login_form_is_visible(page)
    trace_step(trace, "login_form_visibility_checked", visible=login_form_visible)

    if not login_form_visible:
        if before_login_path:
            capture_debug_screenshot(
                page,
                before_login_path,
                trace,
                "login_form_not_detected_screenshot",
            )

        try:
            trace_step(trace, "try_inventory_without_login_form", url=inventory_url())
            page.goto(inventory_url(), wait_until="domcontentloaded", timeout=30_000)
            trace_page_state(trace, "inventory_without_login_form_loaded", page)
            wait_for_inventory_page(page, trace=trace)
            save_session_state(context)
            trace_step(trace, "session_saved", path=SESSION_STATE_PATH)
            return context, page, False, True
        except Exception as exc:
            trace_step(trace, "inventory_without_login_form_failed", error=f"{type(exc).__name__}: {exc}")
            trace_step(trace, "return_to_login_url", url=login_url())
            page.goto(login_url(), wait_until="domcontentloaded", timeout=30_000)
            trace_page_state(trace, "login_page_reloaded", page)

    trace_step(trace, "fill_login_form")
    fill_login_form(page)

    if before_login_path:
        trace_step(trace, "capture_before_login_screenshot", path=before_login_path)
        page.screenshot(path=before_login_path, full_page=True)

    trace_step(trace, "submit_login_form")
    submit_login_form(page)
    wait_for_inventory_after_login_submit(page, trace=trace)
    save_session_state(context)
    trace_step(trace, "session_saved", path=SESSION_STATE_PATH)

    return context, page, False, True


def open_authenticated_inventory_page(browser, before_login_path=None, trace=None):
    saved_session = open_saved_session(browser, trace=trace)

    if saved_session is not None:
        return saved_session

    return open_fresh_session(browser, before_login_path=before_login_path, trace=trace)


def click_availability_filter_button(page):
    result = page.evaluate(
        """
        () => {
          const normalize = (value) => (value || '').replace(/\\s+/g, ' ').trim();
          const isVisible = (element) => {
            const rect = element.getBoundingClientRect();
            const style = window.getComputedStyle(element);
            return (
              rect.width > 0 &&
              rect.height > 0 &&
              style.visibility !== 'hidden' &&
              style.display !== 'none' &&
              style.opacity !== '0'
            );
          };
          const headers = Array.from(document.querySelectorAll('th'));
          const header = headers.find((element) => (
            normalize(element.innerText).includes('Verfügbarkeit') && isVisible(element)
          ));

          if (!header) {
            return { ok: false, error: 'availability_header_not_found' };
          }

          const icons = Array.from(header.querySelectorAll('i'));
          const icon = icons.find((element) => normalize(element.textContent) === 'filter_alt');
          const button = icon ? icon.closest('button') : null;

          if (!button) {
            return { ok: false, error: 'availability_filter_button_not_found' };
          }

          button.click();
          return { ok: true, header_text: normalize(header.innerText) };
        }
        """
    )

    if not result.get("ok"):
        raise RuntimeError(result.get("error") or "availability_filter_button_failed")

    return result


def wait_for_availability_filter_menu(page, trace=None):
    trace_step(trace, "wait_for_availability_filter_menu", timeout_ms=FILTER_TIMEOUT_MS)
    page.wait_for_function(
        """
        () => {
          const fold = (value) => (value || '')
            .normalize('NFD')
            .replace(/[\\u0300-\\u036f]/g, '')
            .replace(/\\s+/g, ' ')
            .trim()
            .toLowerCase();
          const candidates = Array.from(document.querySelectorAll(
            '.q-menu *, .q-position-engine *, .q-checkbox, [role="checkbox"], label, span, div'
          ));

          return candidates.some((element) => {
            if (element.closest('th')) {
              return false;
            }

            if (fold(element.textContent) !== 'verfugbar') {
              return false;
            }

            const rect = element.getBoundingClientRect();
            const style = window.getComputedStyle(element);

            return (
              rect.width > 0 &&
              rect.height > 0 &&
              style.visibility !== 'hidden' &&
              style.display !== 'none' &&
              style.opacity !== '0'
            );
          });
        }
        """,
        timeout=FILTER_TIMEOUT_MS,
    )
    wait_for_next_render_frame(page)
    trace_page_state(trace, "availability_filter_menu_visible", page)

    return {"ok": True, "menu_item": "verfügbar"}


def ensure_available_checkbox_checked(page):
    page.locator('[role="checkbox"][aria-label="verfügbar"]').first.wait_for(
        state="visible",
        timeout=FILTER_TIMEOUT_MS,
    )

    result = page.evaluate(
        """
        () => {
          const checkbox = document.querySelector('[role="checkbox"][aria-label="verfügbar"]');

          if (!checkbox) {
            return { ok: false, error: 'available_checkbox_not_found' };
          }

          const nativeInput = checkbox.querySelector('input[type="checkbox"]');
          const ariaChecked = checkbox.getAttribute('aria-checked');
          const className = String(checkbox.className || '');
          const checked =
            (nativeInput && nativeInput.checked) ||
            ariaChecked === 'true' ||
            className.includes('--truthy') ||
            Boolean(checkbox.querySelector('.q-checkbox__inner--truthy'));

          if (!checked) {
            const box =
              checkbox.querySelector('.q-checkbox__inner') ||
              checkbox.querySelector('.q-checkbox__bg') ||
              checkbox;
            box.click();
          }

          return { ok: true, was_checked: Boolean(checked) };
        }
        """
    )

    if not result.get("ok"):
        raise RuntimeError(result.get("error") or "available_checkbox_failed")

    return result


def get_visible_availability_summary(page):
    return page.evaluate(
        """
        () => {
          const normalize = (value) => (value || '').replace(/\\s+/g, ' ').trim();
          const fold = (value) => normalize(value)
            .normalize('NFD')
            .replace(/[\\u0300-\\u036f]/g, '')
            .toLowerCase();
          const isVisible = (element) => {
            const rect = element.getBoundingClientRect();
            const style = window.getComputedStyle(element);
            return (
              rect.width > 0 &&
              rect.height > 0 &&
              style.visibility !== 'hidden' &&
              style.display !== 'none' &&
              style.opacity !== '0'
            );
          };
          const table = document.querySelector('table.q-table');

          if (!table) {
            return { ok: false, error: 'inventory_table_not_found' };
          }

          const headers = Array.from(table.querySelectorAll('thead th'));
          const availabilityIndex = headers.findIndex((header) => fold(header.innerText).includes('verfugbarkeit'));

          if (availabilityIndex < 0) {
            return { ok: false, error: 'availability_column_not_found' };
          }

          const rows = Array.from(table.querySelectorAll('tbody tr')).filter(isVisible);
          const statuses = rows.map((row, rowIndex) => {
            const cell = row.querySelectorAll('td')[availabilityIndex];
            const text = normalize(cell ? cell.textContent : '');
            const positive = Boolean(cell && cell.querySelector('.text-positive'));
            const negative = Boolean(
              cell && cell.querySelector('.text-negative, .text-red, .text-red-7, .text-red-8')
            );

            return {
              row_index: rowIndex + 1,
              text,
              available: positive ? true : (negative ? false : null),
            };
          });
          const unavailable = statuses.filter((status) => status.available !== true);

          return {
            ok: true,
            row_count: rows.length,
            available_count: statuses.filter((status) => status.available === true).length,
            unavailable_count: unavailable.length,
            unavailable_sample: unavailable.slice(0, 5),
          };
        }
        """
    )


def wait_for_only_available_visible_rows(page, trace=None):
    trace_step(trace, "wait_for_only_available_visible_rows", timeout_ms=SCRAPE_PAGE_READY_TIMEOUT_MS)
    page.wait_for_function(
        """
        () => {
          const normalize = (value) => (value || '').replace(/\\s+/g, ' ').trim();
          const fold = (value) => normalize(value)
            .normalize('NFD')
            .replace(/[\\u0300-\\u036f]/g, '')
            .toLowerCase();
          const isVisible = (element) => {
            const rect = element.getBoundingClientRect();
            const style = window.getComputedStyle(element);
            return (
              rect.width > 0 &&
              rect.height > 0 &&
              style.visibility !== 'hidden' &&
              style.display !== 'none' &&
              style.opacity !== '0'
            );
          };
          const table = document.querySelector('table.q-table');

          if (!table) {
            return false;
          }

          const headers = Array.from(table.querySelectorAll('thead th'));
          const availabilityIndex = headers.findIndex((header) => fold(header.innerText).includes('verfugbarkeit'));

          if (availabilityIndex < 0) {
            return false;
          }

          const rows = Array.from(table.querySelectorAll('tbody tr')).filter(isVisible);

          if (rows.length === 0) {
            return false;
          }

          return rows.every((row) => {
            const cell = row.querySelectorAll('td')[availabilityIndex];
            return Boolean(cell && cell.querySelector('.text-positive'));
          });
        }
        """,
        timeout=SCRAPE_PAGE_READY_TIMEOUT_MS,
    )
    wait_for_next_render_frame(page)
    summary = get_visible_availability_summary(page)
    trace_step(trace, "only_available_visible_rows_ready", **summary)

    return summary


def apply_available_filter(page, trace=None):
    wait_for_inventory_page(page, trace=trace)
    trace_step(trace, "click_availability_filter")
    click_result = click_availability_filter_button(page)
    wait_for_availability_filter_menu(page, trace=trace)
    trace_step(trace, "ensure_available_checkbox_checked")
    checkbox_result = ensure_available_checkbox_checked(page)
    page.keyboard.press("Escape")
    wait_for_inventory_page(page, trace=trace)
    availability_summary = wait_for_only_available_visible_rows(page, trace=trace)

    return {
        "filter_button": click_result,
        "available_checkbox": checkbox_result,
        "availability_summary": availability_summary,
        "visible_rows": availability_summary.get("row_count"),
    }


def normalize_space(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def parse_decimal(value):
    text = normalize_space(value)

    if not text or text.lower() in {"n.a.", "na", "n/a", "-"} or "<" in text:
        return None

    text = text.replace("€", "").replace(" ", "").replace(".", "").replace(",", ".")
    match = re.search(r"-?\d+(?:\.\d+)?", text)

    if not match:
        return None

    try:
        return Decimal(match.group(0))
    except InvalidOperation:
        return None


def parse_date(value):
    text = normalize_space(value)

    if not text or text.lower() in {"n.a.", "na", "n/a", "-"}:
        return None

    match = re.fullmatch(r"(\d{1,2})\.(\d{1,2})\.(\d{4})", text)
    if not match:
        return None

    day, month, year = map(int, match.groups())
    if year < 1900:
        return None

    try:
        return datetime(year, month, day).date()
    except ValueError:
        return None


def postgres_url_value():
    value = (
        os.environ.get("SUPABASE_DB_URL")
        or os.environ.get("DATABASE_URL")
        or os.environ.get("POSTGRES_URL")
        or ""
    ).strip()

    if value and value.startswith(("postgresql://", "postgres://")) and "sslmode=" not in value:
        separator = "&" if "?" in value else "?"
        value = f"{value}{separator}sslmode=require"

    return value


def postgres_url_configured():
    return bool(postgres_url_value())


def postgres_url():
    value = postgres_url_value()

    if not value:
        raise RuntimeError(
            "Missing SUPABASE_DB_URL. Use the Supabase direct Postgres or pooler URL "
            "with sslmode=require."
        )

    return value


def products_schema():
    return (
        os.environ.get("WAWICAN_PRODUCTS_SCHEMA")
        or os.environ.get("SUPABASE_SCHEMA")
        or "private"
    ).strip() or "private"


def products_table():
    return (os.environ.get("WAWICAN_PRODUCTS_TABLE") or "wawican_products").strip() or "wawican_products"


def products_flat_view():
    return (
        os.environ.get("WAWICAN_PRODUCTS_FLAT_VIEW")
        or f"{products_table()}_flat"
    ).strip() or "wawican_products_flat"


DIRECT_PRODUCT_COLUMNS = [
    "product_name",
    "price_per_g_text",
    "price_per_g",
    "net_purchase_price_per_g_text",
    "net_purchase_price_per_g",
    "availability_status",
    "available",
    "actual_stock_text",
    "actual_stock",
    "virtual_stock_text",
    "virtual_stock",
    "price_calculation_enabled",
    "always_available",
    "remaining_quantity_text",
    "remaining_quantity",
    "cultivar",
    "genetics",
    "dominance",
    "thc",
    "cbd",
    "supplier_reserved_quantity_text",
    "supplier_reserved_quantity",
    "expiry_date_text",
    "expiry_date",
    "expires_at",
    "hidden",
    "do_not_show",
    "page_number",
    "row_index",
    "source_url",
    "scraped_at",
]


def rest_insert_columns():
    value = (os.environ.get("WAWICAN_PRODUCTS_REST_COLUMNS") or "").strip()

    if not value:
        return list(DIRECT_PRODUCT_COLUMNS)

    if value == "*":
        return None

    columns = [column.strip() for column in value.split(",") if column.strip()]

    if columns == ["product_name", "raw_data", "scraped_at"] and not bool_env(
        "WAWICAN_PRODUCTS_ALLOW_RAW_REST_COLUMNS",
        False,
    ):
        return list(DIRECT_PRODUCT_COLUMNS)

    return columns


def products_upsert_on():
    value = (os.environ.get("WAWICAN_PRODUCTS_UPSERT_ON") or "product_name").strip()
    return value or None


def supabase_url():
    return (os.environ.get("SUPABASE_URL") or "").strip().rstrip("/")


def supabase_service_role_key():
    return (os.environ.get("SUPABASE_SERVICE_ROLE_KEY") or "").strip()


def supabase_rest_configured():
    return bool(supabase_url() and supabase_service_role_key())


def supabase_table_url():
    return f"{supabase_url()}/rest/v1/{quote(products_table(), safe='')}"


def supabase_headers():
    service_role_key = required_env("SUPABASE_SERVICE_ROLE_KEY")
    schema_name = products_schema()

    return {
        "apikey": service_role_key,
        "Authorization": f"Bearer {service_role_key}",
        "Accept-Profile": schema_name,
        "Content-Profile": schema_name,
        "Content-Type": "application/json",
    }


def response_preview(response):
    body = response.text

    if len(body) > 1000:
        body = f"{body[:1000]}..."

    return {
        "status_code": response.status_code,
        "ok": response.status_code < 400,
        "body": body,
    }


def missing_column_from_postgrest_response(response):
    try:
        payload = response.json()
    except Exception:
        payload = {}

    message = str(payload.get("message") or response.text or "")
    match = re.search(r"Could not find the '([^']+)' column", message)

    if match:
        return match.group(1)

    match = re.search(r"column\s+\w+\.([A-Za-z_][A-Za-z0-9_]*)\s+does not exist", message)

    if match:
        return match.group(1)

    return None


def validate_identifier(value, label):
    if not re.fullmatch(r"[A-Za-z_][A-Za-z0-9_]*", value or ""):
        raise RuntimeError(f"Invalid {label}: {value!r}")

    return value


def scrape_current_inventory_page(page, page_number=None, trace=None):
    trace_step(trace, "wait_for_inventory_table", timeout_ms=SCRAPE_PAGE_READY_TIMEOUT_MS)
    page.locator("table.q-table tbody").first.wait_for(
        state="visible",
        timeout=SCRAPE_PAGE_READY_TIMEOUT_MS,
    )

    result = page.evaluate(
        """
        () => {
          const normalize = (value) => (value || '').replace(/\\s+/g, ' ').trim();
          const fold = (value) => normalize(value)
            .normalize('NFD')
            .replace(/[\\u0300-\\u036f]/g, '')
            .toLowerCase();
          const isVisible = (element) => {
            if (!element) return false;
            const rect = element.getBoundingClientRect();
            const style = window.getComputedStyle(element);
            return (
              rect.width > 0 &&
              rect.height > 0 &&
              style.visibility !== 'hidden' &&
              style.display !== 'none' &&
              style.opacity !== '0'
            );
          };
          const removeNoise = (element) => {
            const clone = element.cloneNode(true);
            clone.querySelectorAll(
              'button, svg, img, .q-focus-helper, .q-tooltip, .q-menu, .q-position-engine'
            ).forEach((node) => node.remove());
            return normalize(clone.textContent);
          };
          const switchValue = (cell) => {
            const switchElement = cell.querySelector('[role="switch"]');
            if (!switchElement) {
              return null;
            }

            return switchElement.getAttribute('aria-checked') === 'true';
          };
          const mapHeader = (header) => {
            const key = fold(header);
            if (key.includes('blute')) return 'product_name';
            if (key === 'preis pro g') return 'price_per_g_text';
            if (key.includes('netto ek')) return 'net_purchase_price_per_g_text';
            if (key.includes('verfugbarkeit')) return 'availability_status';
            if (key.includes('ist bestand')) return 'actual_stock_text';
            if (key.includes('virtueller bestand')) return 'virtual_stock_text';
            if (key.includes('preisberechnung')) return 'price_calculation_enabled';
            if (key.includes('immer verfugbar')) return 'always_available';
            if (key.includes('restmenge')) return 'remaining_quantity_text';
            if (key.includes('kultivar')) return 'cultivar';
            if (key.includes('genetik')) return 'genetics';
            if (key.includes('dominanz')) return 'dominance';
            if (key === 'thc') return 'thc';
            if (key === 'cbd') return 'cbd';
            if (key.includes('reservierung beim lieferanten')) return 'supplier_reserved_quantity_text';
            if (key.includes('verfall')) return 'expiry_date_text';
            if (key.includes('nicht anzeigen')) return 'hidden';
            return key.replace(/[^a-z0-9]+/g, '_').replace(/^_+|_+$/g, '');
          };

          const table = document.querySelector('table.q-table');
          if (!table) {
            return { ok: false, error: 'inventory_table_not_found' };
          }

          const headers = Array.from(table.querySelectorAll('thead th')).map((header) => ({
            label: removeNoise(header),
            field: mapHeader(removeNoise(header)),
          }));
          const domRows = Array.from(table.querySelectorAll('tbody tr'));
          const visibleRows = domRows.filter(isVisible);
          const rows = visibleRows.map((row, rowIndex) => {
            const record = {
              row_index: rowIndex + 1,
              raw_cells: [],
            };

            Array.from(row.querySelectorAll('td')).forEach((cell, cellIndex) => {
              const header = headers[cellIndex] || {
                label: `column_${cellIndex + 1}`,
                field: `column_${cellIndex + 1}`,
              };
              const text = removeNoise(cell);
              const iconTexts = Array.from(cell.querySelectorAll('i')).map((icon) => normalize(icon.textContent));
              const toggle = switchValue(cell);
              const rawCell = {
                index: cellIndex + 1,
                header: header.label,
                field: header.field,
                text,
                icon_texts: iconTexts,
                switch_value: toggle,
              };

              record.raw_cells.push(rawCell);

              if (header.field === 'availability_status') {
                const hasPositive = Boolean(cell.querySelector('.text-positive'));
                const hasNegative = Boolean(
                  cell.querySelector('.text-negative, .text-red, .text-red-7, .text-red-8')
                );
                record.available = hasPositive ? true : (hasNegative ? false : null);
                record.availability_status = hasPositive ? 'verfügbar' : (hasNegative ? 'nicht verfügbar' : text);
                return;
              }

              if (
                header.field === 'price_calculation_enabled' ||
                header.field === 'always_available' ||
                header.field === 'hidden'
              ) {
                record[header.field] = toggle;
                return;
              }

              record[header.field] = text;
            });

            return record;
          });

          return {
            ok: true,
            headers,
            row_count: rows.length,
            dom_row_count: domRows.length,
            visible_row_count: visibleRows.length,
            rows,
          };
        }
        """
    )

    if not result.get("ok"):
        raise RuntimeError(result.get("error") or "inventory_scrape_failed")

    trace_step(
        trace,
        "scraped_inventory_page",
        page_number=page_number,
        row_count=result.get("row_count", 0),
        dom_row_count=result.get("dom_row_count", 0),
        visible_row_count=result.get("visible_row_count", 0),
    )

    return result


def get_pagination_state(page):
    return page.evaluate(
        """
        () => {
          const normalize = (value) => (value || '').replace(/\\s+/g, ' ').trim();
          const cleanCellText = (element) => {
            if (!element) return '';
            const clone = element.cloneNode(true);
            clone.querySelectorAll('button, svg, img, .q-focus-helper, .q-tooltip, .q-menu, .q-position-engine')
              .forEach((node) => node.remove());
            return normalize(clone.textContent);
          };
          const isVisible = (element) => {
            if (!element) return false;
            const rect = element.getBoundingClientRect();
            const style = window.getComputedStyle(element);
            return (
              rect.width > 0 &&
              rect.height > 0 &&
              style.visibility !== 'hidden' &&
              style.display !== 'none' &&
              style.opacity !== '0'
            );
          };
          const parsePageLabel = (label) => {
            const match = normalize(label).match(/(\\d+)\\s*\\/\\s*(\\d+)/);
            if (!match) {
              return { current: null, total: null };
            }

            return { current: Number(match[1]), total: Number(match[2]) };
          };
          const pagination = document.querySelector('.q-pagination');
          const input = pagination ? pagination.querySelector('input[type="number"], input') : null;
          const label = input
            ? normalize(input.value || input.getAttribute('placeholder') || '')
            : normalize(pagination ? pagination.textContent : '');
          const parsed = parsePageLabel(label);
          const nextButtons = Array.from(document.querySelectorAll('.q-pagination button')).filter((button) => {
            const icon = button.querySelector('i');
            return normalize(icon ? icon.textContent : '') === 'keyboard_arrow_right' && isVisible(button);
          });
          const nextButton = nextButtons[0] || null;
          const nextDisabled = !nextButton || nextButton.disabled ||
            nextButton.getAttribute('aria-disabled') === 'true' ||
            nextButton.classList.contains('disabled');
          const visibleRows = Array.from(document.querySelectorAll('table.q-table tbody tr')).filter(isVisible);
          const visibleProductNames = visibleRows.map((row) => (
            cleanCellText(row.querySelector('td.product-name-body') || row.querySelector('td'))
          )).filter(Boolean);

          return {
            label,
            current_page: parsed.current,
            total_pages: parsed.total,
            has_next: !nextDisabled,
            visible_next_buttons: nextButtons.length,
            visible_row_count: visibleRows.length,
            first_row_text: visibleProductNames[0] || '',
            last_row_text: visibleProductNames[visibleProductNames.length - 1] || '',
            visible_product_signature: visibleProductNames.join('||'),
          };
        }
        """
    )


def click_next_inventory_page(page, before_state, trace=None):
    trace_step(trace, "click_next_inventory_page", before_state=before_state)
    result = page.evaluate(
        """
        (before) => {
          const normalize = (value) => (value || '').replace(/\\s+/g, ' ').trim();
          const isVisible = (element) => {
            if (!element) return false;
            const rect = element.getBoundingClientRect();
            const style = window.getComputedStyle(element);
            return (
              rect.width > 0 &&
              rect.height > 0 &&
              style.visibility !== 'hidden' &&
              style.display !== 'none' &&
              style.opacity !== '0'
            );
          };
          const candidates = Array.from(document.querySelectorAll('.q-pagination button')).filter((button) => {
            const icon = button.querySelector('i');
            return normalize(icon ? icon.textContent : '') === 'keyboard_arrow_right' && isVisible(button);
          });
          const enabledCandidates = candidates.filter((button) => !(
            button.disabled ||
            button.getAttribute('aria-disabled') === 'true' ||
            button.classList.contains('disabled')
          ));
          const nextButton = enabledCandidates[0] || null;

          if (!nextButton) {
            return {
              ok: true,
              clicked: false,
              reason: 'next_button_disabled',
              visible_next_buttons: candidates.length,
              enabled_next_buttons: enabledCandidates.length,
            };
          }

          nextButton.click();
          return {
            ok: true,
            clicked: true,
            expected_page: before.current_page ? before.current_page + 1 : null,
            visible_next_buttons: candidates.length,
            enabled_next_buttons: enabledCandidates.length,
          };
        }
        """,
        arg=before_state,
    )

    if not result.get("ok"):
        raise RuntimeError(result.get("error") or "next_page_click_failed")

    if not result.get("clicked"):
        return result

    page.wait_for_function(
        """
        (before) => {
          const normalize = (value) => (value || '').replace(/\\s+/g, ' ').trim();
          const cleanCellText = (element) => {
            if (!element) return '';
            const clone = element.cloneNode(true);
            clone.querySelectorAll('button, svg, img, .q-focus-helper, .q-tooltip, .q-menu, .q-position-engine')
              .forEach((node) => node.remove());
            return normalize(clone.textContent);
          };
          const isVisible = (element) => {
            if (!element) return false;
            const rect = element.getBoundingClientRect();
            const style = window.getComputedStyle(element);
            return (
              rect.width > 0 &&
              rect.height > 0 &&
              style.visibility !== 'hidden' &&
              style.display !== 'none' &&
              style.opacity !== '0'
            );
          };
          const pagination = document.querySelector('.q-pagination');
          const input = pagination ? pagination.querySelector('input[type="number"], input') : null;
          const label = input
            ? normalize(input.value || input.getAttribute('placeholder') || '')
            : normalize(pagination ? pagination.textContent : '');
          const match = label.match(/(\\d+)\\s*\\/\\s*(\\d+)/);
          const current = match ? Number(match[1]) : null;
          const visibleRows = Array.from(document.querySelectorAll('table.q-table tbody tr')).filter(isVisible);
          const visibleProductNames = visibleRows.map((row) => (
            cleanCellText(row.querySelector('td.product-name-body') || row.querySelector('td'))
          )).filter(Boolean);
          const signature = visibleProductNames.join('||');
          const firstRowText = visibleProductNames[0] || '';
          const expected = before.current_page ? before.current_page + 1 : null;
          const pageChanged = expected
            ? current === expected
            : Boolean(before.current_page && current && current !== before.current_page);
          const signatureChanged = Boolean(
            signature &&
            before.visible_product_signature &&
            signature !== before.visible_product_signature
          );

          if (!pageChanged || !signatureChanged || visibleProductNames.length === 0) {
            window.__wawicanTableReadyProbe = null;
            return false;
          }

          const probeKey = `${current}|${visibleProductNames.length}|${signature}`;
          const previousProbe = window.__wawicanTableReadyProbe || {};

          if (previousProbe.key === probeKey) {
            previousProbe.hits = (previousProbe.hits || 1) + 1;
            window.__wawicanTableReadyProbe = previousProbe;
          } else {
            window.__wawicanTableReadyProbe = { key: probeKey, hits: 1 };
          }

          return window.__wawicanTableReadyProbe.hits >= 2;
        }
        """,
        arg=before_state,
        timeout=SCRAPE_PAGE_READY_TIMEOUT_MS,
    )
    wait_for_inventory_page(page, timeout=SCRAPE_PAGE_READY_TIMEOUT_MS, trace=trace)

    return result


def normalize_scraped_product(raw_product, page_number, source_url, scraped_at):
    product = {
        "product_name": normalize_space(raw_product.get("product_name")),
        "price_per_g_text": normalize_space(raw_product.get("price_per_g_text")),
        "price_per_g": parse_decimal(raw_product.get("price_per_g_text")),
        "net_purchase_price_per_g_text": normalize_space(raw_product.get("net_purchase_price_per_g_text")),
        "net_purchase_price_per_g": parse_decimal(raw_product.get("net_purchase_price_per_g_text")),
        "availability_status": normalize_space(raw_product.get("availability_status")),
        "available": raw_product.get("available"),
        "actual_stock_text": normalize_space(raw_product.get("actual_stock_text")),
        "actual_stock": parse_decimal(raw_product.get("actual_stock_text")),
        "virtual_stock_text": normalize_space(raw_product.get("virtual_stock_text")),
        "virtual_stock": parse_decimal(raw_product.get("virtual_stock_text")),
        "price_calculation_enabled": raw_product.get("price_calculation_enabled"),
        "always_available": raw_product.get("always_available"),
        "remaining_quantity_text": normalize_space(raw_product.get("remaining_quantity_text")),
        "remaining_quantity": parse_decimal(raw_product.get("remaining_quantity_text")),
        "cultivar": normalize_space(raw_product.get("cultivar")),
        "genetics": normalize_space(raw_product.get("genetics")),
        "dominance": normalize_space(raw_product.get("dominance")),
        "thc": normalize_space(raw_product.get("thc")),
        "cbd": normalize_space(raw_product.get("cbd")),
        "supplier_reserved_quantity_text": normalize_space(raw_product.get("supplier_reserved_quantity_text")),
        "supplier_reserved_quantity": parse_decimal(raw_product.get("supplier_reserved_quantity_text")),
        "expiry_date_text": normalize_space(raw_product.get("expiry_date_text")),
        "expiry_date": parse_date(raw_product.get("expiry_date_text")),
        "expires_at": parse_date(raw_product.get("expiry_date_text")),
        "hidden": raw_product.get("hidden"),
        "do_not_show": raw_product.get("hidden"),
        "page_number": page_number,
        "row_index": raw_product.get("row_index"),
        "source_url": source_url,
        "scraped_at": scraped_at,
        "raw_data": raw_product,
    }

    return product


def raw_product_is_available(raw_product):
    status = normalize_space(raw_product.get("availability_status")).lower()
    folded_status = (
        status
        .replace("ü", "u")
        .replace("ä", "a")
        .replace("ö", "o")
        .replace("ß", "ss")
    )

    return raw_product.get("available") is True or folded_status == "verfugbar"


def normalize_products_for_db(raw_products, source_url, scraped_at):
    return [
        normalize_scraped_product(raw_product, raw_product.get("page_number"), source_url, scraped_at)
        for raw_product in raw_products
        if normalize_space(raw_product.get("product_name")) and raw_product_is_available(raw_product)
    ]


def product_name_key(product):
    return normalize_space(product.get("product_name")).casefold()


def dedupe_products_by_name(products):
    deduped = []
    index_by_name = {}
    duplicate_samples = []
    duplicate_count = 0

    for product in products:
        key = product_name_key(product)

        if not key:
            continue

        if key in index_by_name:
            duplicate_count += 1

            if len(duplicate_samples) < 10:
                original = deduped[index_by_name[key]]
                duplicate_samples.append(
                    {
                        "product_name": product.get("product_name"),
                        "first_page": original.get("page_number"),
                        "duplicate_page": product.get("page_number"),
                    }
                )
            continue

        index_by_name[key] = len(deduped)
        deduped.append(product)

    return deduped, {
        "input_rows": len(products),
        "unique_rows": len(deduped),
        "duplicate_rows": duplicate_count,
        "duplicate_samples": duplicate_samples,
    }


def get_db_columns(cursor, schema_name, table_name):
    cursor.execute(
        """
        select column_name, data_type, udt_name
        from information_schema.columns
        where table_schema = %s
          and table_name = %s
        order by ordinal_position
        """,
        (schema_name, table_name),
    )
    columns = cursor.fetchall()

    if not columns:
        raise RuntimeError(f"Table {schema_name}.{table_name} was not found or has no columns.")

    return {
        row[0]: {
            "data_type": row[1],
            "udt_name": row[2],
        }
        for row in columns
    }


def value_for_column(column_name, column_info, product, jsonb_wrapper):
    value = product.get(column_name)
    data_type = column_info.get("data_type")
    udt_name = column_info.get("udt_name")

    if value is None:
        return None

    if data_type in {"json", "jsonb"} or udt_name in {"json", "jsonb"}:
        return jsonb_wrapper(value)

    if data_type == "boolean":
        if isinstance(value, bool):
            return value
        if isinstance(value, str):
            return value.strip().lower() in {"1", "true", "yes", "on", "verfügbar", "verfugbar"}
        return bool(value)

    if data_type in {"integer", "bigint", "smallint"}:
        parsed = parse_decimal(value)
        return int(parsed) if parsed is not None else None

    if data_type in {"numeric", "decimal", "real", "double precision"}:
        return parse_decimal(value) if not isinstance(value, Decimal) else value

    if data_type == "date":
        return parse_date(value) if isinstance(value, str) else value

    if data_type.startswith("timestamp"):
        return value

    if data_type in {"text", "character varying", "character"}:
        if isinstance(value, (datetime,)):
            return value.isoformat()
        if hasattr(value, "isoformat") and not isinstance(value, str):
            return value.isoformat()
        return str(value)

    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)

    if isinstance(value, Decimal):
        return str(value)

    return value


def json_ready(value):
    if isinstance(value, Decimal):
        return int(value) if value == value.to_integral_value() else float(value)

    if isinstance(value, datetime):
        return value.isoformat()

    if hasattr(value, "isoformat") and not isinstance(value, str):
        return value.isoformat()

    if isinstance(value, list):
        return [json_ready(item) for item in value]

    if isinstance(value, dict):
        return {key: json_ready(item) for key, item in value.items()}

    return value


def product_payload_for_rest(product):
    return {key: json_ready(value) for key, value in product.items()}


def chunks(items, size):
    for index in range(0, len(items), size):
        yield items[index : index + size]


def validate_rest_payload_columns(column_names, trace=None):
    import httpx

    accepted_columns = list(column_names)
    skipped_columns = []

    while accepted_columns:
        response = httpx.get(
            supabase_table_url(),
            headers=supabase_headers(),
            params={
                "select": ",".join(accepted_columns),
                "limit": "0",
            },
            timeout=SUPABASE_TIMEOUT_SECONDS,
        )

        if response.status_code < 400:
            return accepted_columns, skipped_columns

        missing_column = missing_column_from_postgrest_response(response)
        if missing_column and missing_column in accepted_columns:
            accepted_columns.remove(missing_column)
            skipped_columns.append(missing_column)
            trace_step(trace, "skip_missing_supabase_column", column=missing_column)
            continue

        raise RuntimeError(f"Supabase schema check failed: {response_preview(response)}")

    return accepted_columns, skipped_columns


def write_products_to_supabase_rest(products, trace=None):
    replace_all = bool_env("WAWICAN_PRODUCTS_REPLACE_ALL", True)
    dry_run = bool_env("WAWICAN_PRODUCTS_DRY_RUN", False)
    schema_name = validate_identifier(products_schema(), "WAWICAN_PRODUCTS_SCHEMA")
    table_name = validate_identifier(products_table(), "WAWICAN_PRODUCTS_TABLE")
    upsert_on = products_upsert_on()

    if upsert_on:
        validate_identifier(upsert_on, "WAWICAN_PRODUCTS_UPSERT_ON")

    if dry_run:
        return {
            "enabled": False,
            "dry_run": True,
            "schema": schema_name,
            "table": table_name,
            "inserted_rows": 0,
            "scraped_rows": len(products),
        }

    import httpx

    trace_step(
        trace,
        "write_products_to_supabase_rest",
        schema=schema_name,
        table=table_name,
        rows=len(products),
        replace_all=replace_all,
        upsert_on=upsert_on,
    )

    payloads = [product_payload_for_rest(product) for product in products]
    insert_columns = []
    skipped_columns = []

    if payloads:
        configured_columns = rest_insert_columns()
        if configured_columns is not None:
            payloads = [
                {column: payload.get(column) for column in configured_columns}
                for payload in payloads
            ]

        insert_columns, skipped_columns = validate_rest_payload_columns(
            list(payloads[0].keys()),
            trace=trace,
        )
        payloads = [
            {column: payload.get(column) for column in insert_columns}
            for payload in payloads
        ]

    deleted_rows = None

    if replace_all:
        response = httpx.delete(
            supabase_table_url(),
            headers={
                **supabase_headers(),
                "Prefer": "return=representation,count=exact",
            },
            params={
                "product_name": "not.is.null",
            },
            timeout=SUPABASE_TIMEOUT_SECONDS,
        )

        if response.status_code >= 400:
            raise RuntimeError(f"Supabase cleanup failed: {response_preview(response)}")

        try:
            deleted_payload = response.json()
            if isinstance(deleted_payload, list):
                deleted_rows = len(deleted_payload)
        except Exception:
            deleted_rows = None

    inserted_rows = 0
    insert_url = supabase_table_url()
    prefer_header = "return=minimal"

    if upsert_on:
        insert_url = f"{insert_url}?on_conflict={quote(upsert_on, safe=',')}"
        prefer_header = "resolution=merge-duplicates,return=minimal"

    for product_chunk in chunks(payloads, 500):
        if not product_chunk:
            continue

        response = httpx.post(
            insert_url,
            headers={
                **supabase_headers(),
                "Prefer": prefer_header,
            },
            json=product_chunk,
            timeout=SUPABASE_TIMEOUT_SECONDS,
        )

        if response.status_code >= 400:
            raise RuntimeError(f"Supabase insert failed: {response_preview(response)}")

        inserted_rows += len(product_chunk)

    return {
        "enabled": True,
        "method": "supabase_rest",
        "schema": schema_name,
        "table": table_name,
        "replace_all": replace_all,
        "deleted_rows": deleted_rows,
        "upsert_on": upsert_on,
        "inserted_rows": inserted_rows,
        "insert_columns": insert_columns,
        "skipped_missing_columns": skipped_columns,
        "supabase_url": supabase_url(),
    }


def write_products_to_postgres(products, trace=None):
    replace_all = bool_env("WAWICAN_PRODUCTS_REPLACE_ALL", True)
    dry_run = bool_env("WAWICAN_PRODUCTS_DRY_RUN", False)
    schema_name = validate_identifier(products_schema(), "WAWICAN_PRODUCTS_SCHEMA")
    table_name = validate_identifier(products_table(), "WAWICAN_PRODUCTS_TABLE")

    if dry_run:
        return {
            "enabled": False,
            "dry_run": True,
            "schema": schema_name,
            "table": table_name,
            "inserted_rows": 0,
            "scraped_rows": len(products),
        }

    from psycopg import connect, sql
    from psycopg.types.json import Jsonb

    trace_step(
        trace,
        "write_products_to_postgres",
        schema=schema_name,
        table=table_name,
        rows=len(products),
        replace_all=replace_all,
    )

    with connect(postgres_url()) as connection:
        with connection.cursor() as cursor:
            columns = get_db_columns(cursor, schema_name, table_name)

            if not products:
                if replace_all:
                    cursor.execute(
                        sql.SQL("delete from {}.{}").format(
                            sql.Identifier(schema_name),
                            sql.Identifier(table_name),
                        )
                    )

                return {
                    "enabled": True,
                    "dry_run": False,
                    "schema": schema_name,
                    "table": table_name,
                    "replace_all": replace_all,
                    "inserted_rows": 0,
                    "insert_columns": [],
                }

            insert_columns = [column for column in columns if any(column in product for product in products)]

            if not insert_columns:
                raise RuntimeError(
                    f"No matching columns found in {schema_name}.{table_name}. "
                    "Add product_name/raw_data columns or run the setup SQL."
                )

            if replace_all:
                cursor.execute(
                    sql.SQL("delete from {}.{}").format(
                        sql.Identifier(schema_name),
                        sql.Identifier(table_name),
                    )
                )

            insert_statement = sql.SQL("insert into {}.{} ({}) values ({})").format(
                sql.Identifier(schema_name),
                sql.Identifier(table_name),
                sql.SQL(", ").join(sql.Identifier(column) for column in insert_columns),
                sql.SQL(", ").join(sql.Placeholder() for _ in insert_columns),
            )
            values = [
                [
                    value_for_column(column, columns[column], product, Jsonb)
                    for column in insert_columns
                ]
                for product in products
            ]

            if values:
                cursor.executemany(insert_statement, values)

    return {
        "enabled": True,
        "method": "postgres",
        "dry_run": False,
        "schema": schema_name,
        "table": table_name,
        "replace_all": replace_all,
        "inserted_rows": len(products),
        "insert_columns": insert_columns,
    }


def write_products_to_database(products, trace=None):
    if supabase_rest_configured():
        return write_products_to_supabase_rest(products, trace=trace)

    return write_products_to_postgres(products, trace=trace)


def stock_report_enabled():
    return bool_env("WAWICAN_STOCK_REPORT_ENABLED", True)


def stock_report_n8n_webhook_url():
    return (os.environ.get("WAWICAN_STOCK_REPORT_N8N_WEBHOOK_URL") or "").strip()


def fetch_previous_products_from_supabase_rest(trace=None):
    import httpx

    select_columns = ",".join(
        [
            "product_name",
            "virtual_stock",
            "virtual_stock_text",
            "actual_stock",
            "actual_stock_text",
            "availability_status",
            "available",
            "cultivar",
            "thc",
            "cbd",
            "price_per_g_text",
            "scraped_at",
        ]
    )
    trace_step(trace, "fetch_previous_products_supabase_rest", table=products_table())
    response = httpx.get(
        supabase_table_url(),
        headers=supabase_headers(),
        params={
            "select": select_columns,
            "product_name": "not.is.null",
            "limit": "5000",
        },
        timeout=SUPABASE_TIMEOUT_SECONDS,
    )

    if response.status_code >= 400:
        raise RuntimeError(f"Supabase previous snapshot fetch failed: {response_preview(response)}")

    payload = response.json()
    if not isinstance(payload, list):
        raise RuntimeError("Supabase previous snapshot response was not a list.")

    return payload


def fetch_previous_products_from_postgres(trace=None):
    from psycopg import connect, sql

    schema_name = validate_identifier(products_schema(), "WAWICAN_PRODUCTS_SCHEMA")
    table_name = validate_identifier(products_table(), "WAWICAN_PRODUCTS_TABLE")
    trace_step(trace, "fetch_previous_products_postgres", schema=schema_name, table=table_name)

    with connect(postgres_url()) as connection:
        with connection.cursor() as cursor:
            columns = get_db_columns(cursor, schema_name, table_name)
            wanted_columns = [
                column
                for column in [
                    "product_name",
                    "virtual_stock",
                    "virtual_stock_text",
                    "actual_stock",
                    "actual_stock_text",
                    "availability_status",
                    "available",
                    "cultivar",
                    "thc",
                    "cbd",
                    "price_per_g_text",
                    "scraped_at",
                ]
                if column in columns
            ]

            if "product_name" not in wanted_columns:
                return []

            cursor.execute(
                sql.SQL("select {} from {}.{} where product_name is not null").format(
                    sql.SQL(", ").join(sql.Identifier(column) for column in wanted_columns),
                    sql.Identifier(schema_name),
                    sql.Identifier(table_name),
                )
            )
            rows = cursor.fetchall()

    return [dict(zip(wanted_columns, row)) for row in rows]


def fetch_previous_products_snapshot(trace=None):
    if not stock_report_enabled():
        return []

    if supabase_rest_configured():
        return fetch_previous_products_from_supabase_rest(trace=trace)

    if postgres_url_configured():
        return fetch_previous_products_from_postgres(trace=trace)

    return []


def decimal_for_compare(value):
    if value is None:
        return None

    if isinstance(value, Decimal):
        return value

    if isinstance(value, bool):
        return None

    if isinstance(value, (int, float)):
        try:
            return Decimal(str(value))
        except InvalidOperation:
            return None

    return parse_decimal(value)


def quantity_display(value, fallback=None):
    parsed = decimal_for_compare(value)

    if parsed is None:
        return normalize_space(fallback if fallback is not None else value)

    if parsed == parsed.to_integral_value():
        return str(int(parsed))

    return str(parsed).replace(".", ",")


def excel_number(value):
    parsed = decimal_for_compare(value)

    if parsed is None:
        return None

    if parsed == parsed.to_integral_value():
        return int(parsed)

    return float(parsed)


def compare_stock_snapshots(previous_products, current_products):
    previous_by_name = {
        product_name_key(product): product
        for product in previous_products
        if product_name_key(product)
    }
    current_by_name = {
        product_name_key(product): product
        for product in current_products
        if product_name_key(product)
    }
    changes = []

    for key in sorted(current_by_name, key=lambda item: normalize_space(current_by_name[item].get("product_name")).casefold()):
        current = current_by_name[key]
        previous = previous_by_name.get(key)
        current_stock = decimal_for_compare(current.get("virtual_stock"))
        previous_stock = decimal_for_compare(previous.get("virtual_stock")) if previous else None
        current_text = quantity_display(current.get("virtual_stock"), current.get("virtual_stock_text"))
        previous_text = quantity_display(
            previous.get("virtual_stock") if previous else None,
            previous.get("virtual_stock_text") if previous else None,
        )
        change_type = "new" if previous is None else "changed"

        if previous is not None:
            if current_stock is not None and previous_stock is not None:
                if current_stock == previous_stock:
                    continue
            elif current_text == previous_text:
                continue

        difference = None
        if current_stock is not None and previous_stock is not None:
            difference = current_stock - previous_stock

        changes.append(
            {
                "change_type": change_type,
                "product_name": current.get("product_name"),
                "previous_virtual_stock": previous_stock,
                "current_virtual_stock": current_stock,
                "previous_virtual_stock_text": previous_text,
                "current_virtual_stock_text": current_text,
                "difference": difference,
                "previous_actual_stock": decimal_for_compare(previous.get("actual_stock")) if previous else None,
                "current_actual_stock": decimal_for_compare(current.get("actual_stock")),
                "previous_actual_stock_text": quantity_display(
                    previous.get("actual_stock") if previous else None,
                    previous.get("actual_stock_text") if previous else None,
                ),
                "current_actual_stock_text": quantity_display(current.get("actual_stock"), current.get("actual_stock_text")),
                "availability_status": current.get("availability_status"),
                "cultivar": current.get("cultivar"),
                "thc": current.get("thc"),
                "cbd": current.get("cbd"),
                "price_per_g_text": current.get("price_per_g_text"),
                "page_number": current.get("page_number"),
                "row_index": current.get("row_index"),
                "previous_scraped_at": previous.get("scraped_at") if previous else None,
                "current_scraped_at": current.get("scraped_at"),
            }
        )

    for key in sorted(set(previous_by_name) - set(current_by_name), key=lambda item: normalize_space(previous_by_name[item].get("product_name")).casefold()):
        previous = previous_by_name[key]
        previous_stock = decimal_for_compare(previous.get("virtual_stock"))
        changes.append(
            {
                "change_type": "removed",
                "product_name": previous.get("product_name"),
                "previous_virtual_stock": previous_stock,
                "current_virtual_stock": None,
                "previous_virtual_stock_text": quantity_display(previous.get("virtual_stock"), previous.get("virtual_stock_text")),
                "current_virtual_stock_text": "",
                "difference": None,
                "previous_actual_stock": decimal_for_compare(previous.get("actual_stock")),
                "current_actual_stock": None,
                "previous_actual_stock_text": quantity_display(previous.get("actual_stock"), previous.get("actual_stock_text")),
                "current_actual_stock_text": "",
                "availability_status": "not in current available scrape",
                "cultivar": previous.get("cultivar"),
                "thc": previous.get("thc"),
                "cbd": previous.get("cbd"),
                "price_per_g_text": previous.get("price_per_g_text"),
                "page_number": None,
                "row_index": None,
                "previous_scraped_at": previous.get("scraped_at"),
                "current_scraped_at": None,
            }
        )

    decreased = [
        change
        for change in changes
        if isinstance(change.get("difference"), Decimal) and change["difference"] < 0
    ]
    increased = [
        change
        for change in changes
        if isinstance(change.get("difference"), Decimal) and change["difference"] > 0
    ]

    return changes, {
        "previous_rows": len(previous_products),
        "current_rows": len(current_products),
        "changed_rows": len(changes),
        "decreased_rows": len(decreased),
        "increased_rows": len(increased),
        "new_rows": sum(1 for change in changes if change.get("change_type") == "new"),
        "removed_rows": sum(1 for change in changes if change.get("change_type") == "removed"),
    }


def auto_width_for_sheet(sheet, max_width=60):
    for column_cells in sheet.columns:
        column_letter = column_cells[0].column_letter
        max_length = 0

        for cell in column_cells:
            value = cell.value
            if value is None:
                continue
            max_length = max(max_length, len(str(value)))

        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), max_width)


def parse_datetime_for_excel(value):
    if not value:
        return None
    if isinstance(value, datetime):
        date_value = value
    elif isinstance(value, str):
        normalized = value.strip()
        if not normalized:
            return None
        if normalized.endswith("Z"):
            normalized = f"{normalized[:-1]}+00:00"
        try:
            date_value = datetime.fromisoformat(normalized)
        except ValueError:
            return None
    else:
        return None

    if date_value.tzinfo is None:
        date_value = date_value.replace(tzinfo=timezone.utc)

    return date_value.astimezone(GERMAN_TIMEZONE)


def german_excel_datetime(value):
    date_value = parse_datetime_for_excel(value)
    if not date_value:
        return ""

    return date_value.strftime("%d.%m.%Y %H:%M:%S")


def german_excel_datetime_range(previous_value, current_value):
    previous_text = german_excel_datetime(previous_value)
    current_text = german_excel_datetime(current_value)

    if previous_text and current_text:
        return f"{previous_text} -> {current_text}"
    if current_text:
        return f"neu -> {current_text}"
    if previous_text:
        return f"{previous_text} -> entfernt"
    return ""


def create_stock_change_excel_report(changes, current_products, previous_count, scraped_at, timestamp, trace=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    os.makedirs(ARTIFACTS_DIR, exist_ok=True)
    report_path = os.path.join(ARTIFACTS_DIR, f"wawican-stock-changes-{timestamp}.xlsx")
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Zusammenfassung"
    changes_sheet = workbook.create_sheet("Mengenänderungen")
    current_sheet = workbook.create_sheet("Aktueller Bestand")

    changed_rows = len(changes)
    decreased_rows = sum(
        1
        for change in changes
        if isinstance(change.get("difference"), Decimal) and change["difference"] < 0
    )
    increased_rows = sum(
        1
        for change in changes
        if isinstance(change.get("difference"), Decimal) and change["difference"] > 0
    )
    new_rows = sum(1 for change in changes if change.get("change_type") == "new")
    removed_rows = sum(1 for change in changes if change.get("change_type") == "removed")

    summary_rows = [
        ["Wawican Mengenänderungen", ""],
        ["Erstellt am", german_excel_datetime(scraped_at)],
        ["Zeitzone", "Europe/Berlin"],
        ["Vorherige Produkte", previous_count],
        ["Aktuelle Produkte", len(current_products)],
        ["Zeilen im Report", changed_rows],
        ["Bestand gesunken", decreased_rows],
        ["Bestand gestiegen", increased_rows],
        ["Neue Produkte", new_rows],
        ["Nicht mehr im verfügbaren Filter", removed_rows],
        ["Verglichene Spalte", "Virtueller Bestand"],
    ]
    summary_sheet.append(summary_rows[0])
    summary_sheet["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    summary_sheet["A1"].fill = PatternFill("solid", fgColor="1F4E5F")
    summary_sheet["B1"].fill = PatternFill("solid", fgColor="1F4E5F")
    for row in summary_rows[1:]:
        summary_sheet.append(row)
    summary_sheet.column_dimensions["A"].width = 34
    summary_sheet.column_dimensions["B"].width = 34

    headers = [
        "Produkt",
        "Typ",
        "Menge alt",
        "Menge neu",
        "Änderung (neu - alt)",
        "Ist Bestand alt",
        "Ist Bestand neu",
        "Verfügbarkeit",
        "Kultivar",
        "THC",
        "CBD",
        "Preis pro g",
        "Seite",
        "Zeile",
        "Geändert von -> bis",
        "Vorheriger Lauf (DE)",
        "Aktueller Lauf (DE)",
    ]
    changes_sheet.append(headers)
    if changes:
        for change in changes:
            changes_sheet.append(
                [
                    change.get("product_name"),
                    change.get("change_type"),
                    excel_number(change.get("previous_virtual_stock"))
                    if change.get("previous_virtual_stock") is not None
                    else change.get("previous_virtual_stock_text"),
                    excel_number(change.get("current_virtual_stock"))
                    if change.get("current_virtual_stock") is not None
                    else change.get("current_virtual_stock_text"),
                    excel_number(change.get("difference")) if change.get("difference") is not None else "",
                    excel_number(change.get("previous_actual_stock"))
                    if change.get("previous_actual_stock") is not None
                    else change.get("previous_actual_stock_text"),
                    excel_number(change.get("current_actual_stock"))
                    if change.get("current_actual_stock") is not None
                    else change.get("current_actual_stock_text"),
                    change.get("availability_status"),
                    change.get("cultivar"),
                    change.get("thc"),
                    change.get("cbd"),
                    change.get("price_per_g_text"),
                    change.get("page_number"),
                    change.get("row_index"),
                    german_excel_datetime_range(
                        change.get("previous_scraped_at"),
                        change.get("current_scraped_at"),
                    ),
                    german_excel_datetime(change.get("previous_scraped_at")),
                    german_excel_datetime(change.get("current_scraped_at")),
                ]
            )
    else:
        changes_sheet.append(["Keine Mengenänderungen seit dem letzten Lauf.", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""])

    current_headers = [
        "Produkt",
        "Virtueller Bestand",
        "Ist Bestand",
        "Verfügbarkeit",
        "Kultivar",
        "THC",
        "CBD",
        "Preis pro g",
        "Seite",
        "Zeile",
        "Lauf (DE)",
    ]
    current_sheet.append(current_headers)
    for product in sorted(current_products, key=lambda item: normalize_space(item.get("product_name")).casefold()):
        current_sheet.append(
            [
                product.get("product_name"),
                excel_number(product.get("virtual_stock")) if product.get("virtual_stock") is not None else product.get("virtual_stock_text"),
                excel_number(product.get("actual_stock")) if product.get("actual_stock") is not None else product.get("actual_stock_text"),
                product.get("availability_status"),
                product.get("cultivar"),
                product.get("thc"),
                product.get("cbd"),
                product.get("price_per_g_text"),
                product.get("page_number"),
                product.get("row_index"),
                german_excel_datetime(product.get("scraped_at")),
            ]
        )

    header_fill = PatternFill("solid", fgColor="1F4E5F")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(bottom=Side(style="thin", color="B7B7B7"))
    decrease_fill = PatternFill("solid", fgColor="FCE4D6")
    increase_fill = PatternFill("solid", fgColor="E2F0D9")
    new_fill = PatternFill("solid", fgColor="DDEBF7")
    removed_fill = PatternFill("solid", fgColor="E7E6E6")

    for sheet in [changes_sheet, current_sheet]:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.border = thin_border
            if sheet.title == "Mengenänderungen":
                difference = row[4].value
                change_type = row[1].value
                fill = None
                if isinstance(difference, (int, float)) and difference < 0:
                    fill = decrease_fill
                elif isinstance(difference, (int, float)) and difference > 0:
                    fill = increase_fill
                elif change_type == "new":
                    fill = new_fill
                elif change_type == "removed":
                    fill = removed_fill
                if fill:
                    for cell in row:
                        cell.fill = fill
        auto_width_for_sheet(sheet)

    for sheet in [changes_sheet, current_sheet]:
        for column_index in [3, 4, 5, 6, 7] if sheet.title == "Mengenänderungen" else [2, 3]:
            column_letter = get_column_letter(column_index)
            for cell in sheet[column_letter][1:]:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '0.##'

    workbook.save(report_path)
    trace_step(
        trace,
        "created_stock_change_excel_report",
        path=report_path,
        changed_rows=changed_rows,
        current_rows=len(current_products),
        previous_rows=previous_count,
    )

    return report_path


def send_stock_report_to_n8n(report_path, metadata):
    webhook_url = stock_report_n8n_webhook_url()

    if not webhook_url:
        return {
            "sent_to_n8n": False,
            "n8n_skipped_reason": "WAWICAN_STOCK_REPORT_N8N_WEBHOOK_URL is not configured",
        }

    import httpx

    filename = os.path.basename(report_path)
    content_type = (
        mimetypes.guess_type(filename)[0]
        or "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    with open(report_path, "rb") as file_handle:
        response = httpx.post(
            webhook_url,
            data={key: "" if value is None else str(value) for key, value in metadata.items()},
            files={"file": (filename, file_handle, content_type)},
            timeout=int_env("WAWICAN_STOCK_REPORT_N8N_TIMEOUT_SECONDS", 60),
        )

    if response.status_code >= 400:
        raise RuntimeError(f"n8n stock report upload failed: {response_preview(response)}")

    return {
        "sent_to_n8n": True,
        "n8n_status_code": response.status_code,
    }


def build_stock_report(previous_products, current_products, scraped_at, timestamp, base_url, trace=None):
    if not stock_report_enabled():
        return {
            "enabled": False,
            "reason": "WAWICAN_STOCK_REPORT_ENABLED=false",
        }

    changes, summary = compare_stock_snapshots(previous_products, current_products)
    report_path = create_stock_change_excel_report(
        changes,
        current_products,
        len(previous_products),
        scraped_at,
        timestamp,
        trace=trace,
    )
    n8n_result = send_stock_report_to_n8n(
        report_path,
        {
            "service": SERVICE_NAME,
            "event": "wawican_stock_change_report",
            "run_id": timestamp,
            "changed_rows": summary.get("changed_rows"),
            "current_rows": summary.get("current_rows"),
            "previous_rows": summary.get("previous_rows"),
            "filename": os.path.basename(report_path),
        },
    )

    return {
        "enabled": True,
        **summary,
        "filename": os.path.basename(report_path),
        "path": report_path,
        **report_response_links(base_url, report_path),
        **n8n_result,
    }


def products_flat_view_sql_text(schema_name=None, table_name=None, view_name=None):
    schema_name = validate_identifier(schema_name or products_schema(), "WAWICAN_PRODUCTS_SCHEMA")
    table_name = validate_identifier(table_name or products_table(), "WAWICAN_PRODUCTS_TABLE")
    view_name = validate_identifier(view_name or products_flat_view(), "WAWICAN_PRODUCTS_FLAT_VIEW")

    return f"""
create or replace view {schema_name}.{view_name} as
select
  product_name,
  raw_data->>'price_per_g_text' as price_per_g,
  raw_data->>'net_purchase_price_per_g_text' as net_purchase_price_per_g,
  raw_data->>'availability_status' as availability_status,
  raw_data->>'actual_stock_text' as actual_stock,
  raw_data->>'virtual_stock_text' as virtual_stock,
  raw_data->>'price_calculation_enabled' as price_calculation_enabled,
  raw_data->>'always_available' as always_available,
  raw_data->>'remaining_quantity_text' as remaining_quantity,
  raw_data->>'cultivar' as cultivar,
  raw_data->>'genetics' as genetics,
  raw_data->>'dominance' as dominance,
  raw_data->>'thc' as thc,
  raw_data->>'cbd' as cbd,
  raw_data->>'supplier_reserved_quantity_text' as supplier_reserved_quantity,
  raw_data->>'expiry_date_text' as expiry_date,
  raw_data->>'hidden' as hidden,
  raw_data->>'page_number' as page_number,
  raw_data->>'row_index' as row_index,
  raw_data->'raw_cells' as raw_cells,
  scraped_at,
  raw_data
from {schema_name}.{table_name};

revoke all on {schema_name}.{view_name} from public, anon, authenticated;
grant select on {schema_name}.{view_name} to service_role;
notify pgrst, 'reload schema';
""".strip()


def ensure_products_flat_view(trace=None):
    enabled = bool_env("WAWICAN_PRODUCTS_FLAT_VIEW_ENABLED", False)
    schema_name = validate_identifier(products_schema(), "WAWICAN_PRODUCTS_SCHEMA")
    table_name = validate_identifier(products_table(), "WAWICAN_PRODUCTS_TABLE")
    view_name = validate_identifier(products_flat_view(), "WAWICAN_PRODUCTS_FLAT_VIEW")

    if not enabled:
        return {
            "enabled": False,
            "created": False,
            "schema": schema_name,
            "view": view_name,
        }

    sql_text = products_flat_view_sql_text(schema_name, table_name, view_name)

    if not postgres_url_configured():
        return {
            "enabled": True,
            "created": False,
            "reason": "SUPABASE_DB_URL is required to create database views automatically. Supabase REST cannot run create view.",
            "schema": schema_name,
            "view": view_name,
            "sql": sql_text,
        }

    from psycopg import connect, sql

    trace_step(
        trace,
        "ensure_products_flat_view",
        schema=schema_name,
        table=table_name,
        view=view_name,
    )

    with connect(postgres_url()) as connection:
        with connection.cursor() as cursor:
            cursor.execute(
                sql.SQL(
                    """
                    create or replace view {}.{} as
                    select
                      product_name,
                      raw_data->>'price_per_g_text' as price_per_g,
                      raw_data->>'net_purchase_price_per_g_text' as net_purchase_price_per_g,
                      raw_data->>'availability_status' as availability_status,
                      raw_data->>'actual_stock_text' as actual_stock,
                      raw_data->>'virtual_stock_text' as virtual_stock,
                      raw_data->>'price_calculation_enabled' as price_calculation_enabled,
                      raw_data->>'always_available' as always_available,
                      raw_data->>'remaining_quantity_text' as remaining_quantity,
                      raw_data->>'cultivar' as cultivar,
                      raw_data->>'genetics' as genetics,
                      raw_data->>'dominance' as dominance,
                      raw_data->>'thc' as thc,
                      raw_data->>'cbd' as cbd,
                      raw_data->>'supplier_reserved_quantity_text' as supplier_reserved_quantity,
                      raw_data->>'expiry_date_text' as expiry_date,
                      raw_data->>'hidden' as hidden,
                      raw_data->>'page_number' as page_number,
                      raw_data->>'row_index' as row_index,
                      raw_data->'raw_cells' as raw_cells,
                      scraped_at,
                      raw_data
                    from {}.{}
                    """
                ).format(
                    sql.Identifier(schema_name),
                    sql.Identifier(view_name),
                    sql.Identifier(schema_name),
                    sql.Identifier(table_name),
                )
            )
            cursor.execute(
                sql.SQL("revoke all on {}.{} from public, anon, authenticated").format(
                    sql.Identifier(schema_name),
                    sql.Identifier(view_name),
                )
            )
            cursor.execute(
                sql.SQL("grant select on {}.{} to service_role").format(
                    sql.Identifier(schema_name),
                    sql.Identifier(view_name),
                )
            )
            cursor.execute("notify pgrst, 'reload schema'")

    return {
        "enabled": True,
        "created": True,
        "schema": schema_name,
        "view": view_name,
    }


def scrape_all_available_products(base_url, trace=None):
    os.makedirs(ARTIFACTS_DIR, exist_ok=True)
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    before_login_path = os.path.join(ARTIFACTS_DIR, f"wawican-scrape-before-login-{timestamp}.png")
    screenshot_path = os.path.join(ARTIFACTS_DIR, f"wawican-scrape-products-{timestamp}.png")
    scraped_at = datetime.now(timezone.utc)

    trace_step(trace, "start_browser", headless=bool_env("WAWICAN_HEADLESS", True))
    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=bool_env("WAWICAN_HEADLESS", True))
        context = None

        try:
            context, page, reused_session, inventory_ready = open_authenticated_inventory_page(
                browser,
                before_login_path=before_login_path,
                trace=trace,
            )
            filter_result = apply_available_filter(page, trace=trace)

            raw_products = []
            pages = []

            for page_loop_index in range(1, SCRAPE_MAX_PAGES + 1):
                pagination = get_pagination_state(page)
                page_number = pagination.get("current_page") or page_loop_index
                page_result = scrape_current_inventory_page(
                    page,
                    page_number=page_number,
                    trace=trace,
                )
                page_rows = page_result.get("rows", [])
                availability_summary = get_visible_availability_summary(page)

                if availability_summary.get("unavailable_count"):
                    raise RuntimeError(
                        "Availability filter is not fully applied. "
                        f"Page {page_number} still has {availability_summary.get('unavailable_count')} visible unavailable rows."
                    )

                for row in page_rows:
                    row["page_number"] = page_number
                    raw_products.append(row)

                pages.append(
                    {
                        "page": page_number,
                        "label": pagination.get("label"),
                        "rows": len(page_rows),
                        "available_rows": availability_summary.get("available_count"),
                        "unavailable_rows": availability_summary.get("unavailable_count"),
                        "first_product": normalize_space(page_rows[0].get("product_name")) if page_rows else None,
                        "last_product": normalize_space(page_rows[-1].get("product_name")) if page_rows else None,
                    }
                )

                pagination = get_pagination_state(page)
                current_page = pagination.get("current_page")
                total_pages = pagination.get("total_pages")
                has_next = pagination.get("has_next")

                if not has_next or (current_page and total_pages and current_page >= total_pages):
                    break

                click_next_inventory_page(page, pagination, trace=trace)
                wait_for_only_available_visible_rows(page, trace=trace)
            else:
                raise RuntimeError(f"Stopped after WAWICAN_SCRAPE_MAX_PAGES={SCRAPE_MAX_PAGES}")

            normalized_products = normalize_products_for_db(raw_products, page.url, scraped_at)
            products, dedupe_result = dedupe_products_by_name(normalized_products)
            skipped_non_available_rows = len(
                [
                    row
                    for row in raw_products
                    if normalize_space(row.get("product_name")) and not raw_product_is_available(row)
                ]
            )
            trace_step(
                trace,
                "normalized_available_products",
                raw_rows=len(raw_products),
                available_products=len(normalized_products),
                unique_products=len(products),
                skipped_non_available_rows=skipped_non_available_rows,
                duplicate_available_rows=dedupe_result.get("duplicate_rows"),
            )
            previous_products = fetch_previous_products_snapshot(trace=trace)
            stock_report_result = build_stock_report(
                previous_products,
                products,
                scraped_at,
                timestamp,
                base_url,
                trace=trace,
            )
            db_result = write_products_to_database(products, trace=trace)
            flat_view_result = ensure_products_flat_view(trace=trace)
            screenshot_result = capture_screenshot_now(
                page,
                screenshot_path,
                trace=trace,
                trigger="scrape_finished",
            )

            return {
                "ok": True,
                "current_url": page.url,
                "page_title": page.title(),
                "reused_session": reused_session,
                "inventory_ready": inventory_ready,
                "session_state_path": SESSION_STATE_PATH,
                "before_login_path": None if reused_session else before_login_path,
                "filter": filter_result,
                "pages_scraped": pages,
                "products_scraped": len(products),
                "available_products_seen": len(normalized_products),
                "dedupe": dedupe_result,
                "raw_rows_seen": len(raw_products),
                "skipped_non_available_rows": skipped_non_available_rows,
                "stock_report": stock_report_result,
                "database": db_result,
                "flat_view": flat_view_result,
                **screenshot_result,
                **screenshot_response_links(base_url, screenshot_path),
            }
        finally:
            if context:
                context.close()
            browser.close()


def open_availability_filter_menu(page, trace=None):
    wait_for_inventory_page(page, trace=trace)
    trace_step(trace, "click_availability_filter")
    click_result = click_availability_filter_button(page)
    menu_result = wait_for_availability_filter_menu(page, trace=trace)

    return {
        "filter_button": click_result,
        "filter_menu": menu_result,
    }


def login_only(base_url, trace=None):
    os.makedirs(ARTIFACTS_DIR, exist_ok=True)
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    before_login_path = os.path.join(ARTIFACTS_DIR, f"wawican-before-login-{timestamp}.png")
    after_login_path = os.path.join(ARTIFACTS_DIR, f"wawican-after-login-{timestamp}.png")

    trace_step(trace, "start_browser", headless=bool_env("WAWICAN_HEADLESS", True))
    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=bool_env("WAWICAN_HEADLESS", True))
        context = None

        try:
            context, page, reused_session, inventory_ready = open_authenticated_inventory_page(
                browser,
                before_login_path=before_login_path,
                trace=trace,
            )
            screenshot_result = capture_screenshot_now(page, after_login_path, trace=trace)

            return {
                "ok": True,
                "current_url": page.url,
                "page_title": page.title(),
                "reused_session": reused_session,
                "inventory_ready": inventory_ready,
                "session_state_path": SESSION_STATE_PATH,
                "session_state_exists": os.path.exists(SESSION_STATE_PATH),
                "before_login_path": None if reused_session else before_login_path,
                "after_login_path": after_login_path,
                **screenshot_result,
                **screenshot_response_links(base_url, after_login_path),
            }
        finally:
            if context:
                context.close()
            browser.close()


def login_and_filter_available(base_url, trace=None):
    os.makedirs(ARTIFACTS_DIR, exist_ok=True)
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    before_login_path = os.path.join(ARTIFACTS_DIR, f"wawican-filter-before-login-{timestamp}.png")
    screenshot_path = os.path.join(ARTIFACTS_DIR, f"wawican-filter-available-{timestamp}.png")

    trace_step(trace, "start_browser", headless=bool_env("WAWICAN_HEADLESS", True))
    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=bool_env("WAWICAN_HEADLESS", True))
        context = None

        try:
            context, page, reused_session, inventory_ready = open_authenticated_inventory_page(
                browser,
                before_login_path=before_login_path,
                trace=trace,
            )
            filter_result = apply_available_filter(page, trace=trace)
            screenshot_result = capture_screenshot_now(page, screenshot_path, trace=trace)

            return {
                "ok": True,
                "current_url": page.url,
                "page_title": page.title(),
                "reused_session": reused_session,
                "inventory_ready": inventory_ready,
                "session_state_path": SESSION_STATE_PATH,
                "before_login_path": None if reused_session else before_login_path,
                **filter_result,
                **screenshot_result,
                **screenshot_response_links(base_url, screenshot_path),
            }
        finally:
            if context:
                context.close()
            browser.close()


def login_and_open_availability_filter(base_url, trace=None):
    os.makedirs(ARTIFACTS_DIR, exist_ok=True)
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    before_login_path = os.path.join(ARTIFACTS_DIR, f"wawican-open-filter-before-login-{timestamp}.png")
    screenshot_path = os.path.join(ARTIFACTS_DIR, f"wawican-open-availability-filter-{timestamp}.png")

    trace_step(trace, "start_browser", headless=bool_env("WAWICAN_HEADLESS", True))
    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=bool_env("WAWICAN_HEADLESS", True))
        context = None

        try:
            context, page, reused_session, inventory_ready = open_authenticated_inventory_page(
                browser,
                before_login_path=before_login_path,
                trace=trace,
            )
            filter_result = open_availability_filter_menu(page, trace=trace)
            screenshot_result = capture_screenshot_now(
                page,
                screenshot_path,
                trace=trace,
                trigger="availability_filter_menu_visible",
            )

            return {
                "ok": True,
                "current_url": page.url,
                "page_title": page.title(),
                "reused_session": reused_session,
                "inventory_ready": inventory_ready,
                "session_state_path": SESSION_STATE_PATH,
                "before_login_path": None if reused_session else before_login_path,
                **filter_result,
                **screenshot_result,
                **screenshot_response_links(base_url, screenshot_path),
            }
        finally:
            if context:
                context.close()
            browser.close()


def json_safe(value):
    if value is None or isinstance(value, (str, int, float, bool)):
        return value

    if isinstance(value, (list, tuple)):
        return [json_safe(item) for item in value]

    if isinstance(value, dict):
        return {str(key): json_safe(item) for key, item in value.items()}

    return str(value)


def prune_jobs_locked():
    if len(JOBS) <= MAX_STORED_JOBS:
        return

    sorted_job_ids = sorted(
        JOBS,
        key=lambda job_id: JOBS[job_id].get("created_at", ""),
    )
    for job_id in sorted_job_ids[: len(JOBS) - MAX_STORED_JOBS]:
        JOBS.pop(job_id, None)


def create_job(name):
    job_id = uuid.uuid4().hex
    job = {
        "ok": True,
        "job_id": job_id,
        "name": name,
        "status": "queued",
        "created_at": utc_now_iso(),
        "started_at": None,
        "finished_at": None,
        "last_step": None,
        "steps": [],
    }

    with JOB_LOCK:
        JOBS[job_id] = job
        prune_jobs_locked()

    return job


def append_job_step(job_id, name, **fields):
    step = {
        "name": name,
        "at": utc_now_iso(),
        **{key: json_safe(value) for key, value in fields.items()},
    }

    with JOB_LOCK:
        job = JOBS.get(job_id)
        if not job:
            return

        job.setdefault("steps", []).append(step)
        job["last_step"] = name


def update_job(job_id, **fields):
    with JOB_LOCK:
        job = JOBS.get(job_id)
        if not job:
            return

        job.update({key: json_safe(value) for key, value in fields.items()})


def get_job_snapshot(job_id):
    with JOB_LOCK:
        job = JOBS.get(job_id)
        if not job:
            return None

        return json_safe(job)


def list_job_snapshots():
    with JOB_LOCK:
        jobs = sorted(
            JOBS.values(),
            key=lambda job: job.get("created_at", ""),
            reverse=True,
        )
        return [json_safe(job) for job in jobs]


def run_background_job(job_id, base_url, work):
    update_job(job_id, status="running", started_at=utc_now_iso())
    append_job_step(job_id, "job_started")

    try:
        result = work(base_url, trace=lambda name, **fields: append_job_step(job_id, name, **fields))
        update_job(
            job_id,
            status="done",
            finished_at=utc_now_iso(),
            result=result,
        )
        append_job_step(job_id, "job_done")
    except Exception as exc:
        update_job(
            job_id,
            ok=False,
            status="error",
            finished_at=utc_now_iso(),
            error=f"{type(exc).__name__}: {exc}",
            traceback=traceback.format_exc().splitlines()[-12:],
        )
        append_job_step(job_id, "job_error", error=f"{type(exc).__name__}: {exc}")


def start_background_job(request, name, work):
    base_url = base_url_from_request(request)
    job = create_job(name)
    thread = threading.Thread(
        target=run_background_job,
        args=(job["job_id"], base_url, work),
        daemon=True,
        name=f"{SERVICE_NAME}-{job['job_id'][:8]}",
    )
    thread.start()

    return JSONResponse(
        status_code=202,
        content={
            "ok": True,
            "job_id": job["job_id"],
            "status": "queued",
            "status_url": f"{base_url}/jobs/{job['job_id']}",
        },
    )


@app.get("/health")
def health():
    with JOB_LOCK:
        running_jobs = sum(1 for job in JOBS.values() if job.get("status") in {"queued", "running"})

    return {
        "ok": True,
        "service": SERVICE_NAME,
        "uptime_seconds": round(time.time() - STARTED_AT, 3),
        "inventory_url_configured": bool((os.environ.get("WAWICAN_INVENTORY_URL") or "").strip()),
        "login_url_configured": bool((os.environ.get("WAWICAN_LOGIN_URL") or "").strip()),
        "session_state_path": SESSION_STATE_PATH,
        "session_state_exists": os.path.exists(SESSION_STATE_PATH),
        "screenshot_wait_ms": screenshot_wait_ms(),
        "screenshot_mode": "capture_immediately_when_inventory_ready_visible",
        "inventory_ready_timeout_ms": INVENTORY_READY_TIMEOUT_MS,
        "post_login_ready_timeout_ms": POST_LOGIN_READY_TIMEOUT_MS,
        "scrape_max_pages": SCRAPE_MAX_PAGES,
        "supabase_rest_configured": supabase_rest_configured(),
        "database_url_configured": supabase_rest_configured() or bool(
            (
                os.environ.get("SUPABASE_DB_URL")
                or os.environ.get("DATABASE_URL")
                or os.environ.get("POSTGRES_URL")
                or ""
            ).strip()
        ),
        "postgres_url_configured": postgres_url_configured(),
        "products_schema": products_schema(),
        "products_table": products_table(),
        "products_flat_view": products_flat_view(),
        "products_flat_view_enabled": bool_env("WAWICAN_PRODUCTS_FLAT_VIEW_ENABLED", False),
        "products_rest_columns": rest_insert_columns() or "*",
        "products_upsert_on": products_upsert_on(),
        "products_replace_all": bool_env("WAWICAN_PRODUCTS_REPLACE_ALL", True),
        "stock_report_enabled": stock_report_enabled(),
        "stock_report_n8n_configured": bool(stock_report_n8n_webhook_url()),
        "supabase_timeout_seconds": SUPABASE_TIMEOUT_SECONDS,
        "running_jobs": running_jobs,
    }


@app.post("/jobs/login")
def login_job(request: Request):
    try:
        return login_only(base_url_from_request(request))
    except Exception as exc:
        return JSONResponse(
            status_code=500,
            content={"ok": False, "error": f"{type(exc).__name__}: {exc}"},
        )


@app.post("/jobs/login/start")
def login_job_start(request: Request):
    return start_background_job(request, "login", login_only)


@app.post("/jobs/login/run")
def login_job_sync(request: Request):
    try:
        return login_only(base_url_from_request(request))
    except Exception as exc:
        return JSONResponse(
            status_code=500,
            content={"ok": False, "error": f"{type(exc).__name__}: {exc}"},
        )


@app.post("/jobs/filter-available")
def filter_available_job(request: Request):
    try:
        return login_and_filter_available(base_url_from_request(request))
    except Exception as exc:
        return JSONResponse(
            status_code=500,
            content={"ok": False, "error": f"{type(exc).__name__}: {exc}"},
        )


@app.post("/jobs/filter-available/start")
def filter_available_job_start(request: Request):
    return start_background_job(request, "filter-available", login_and_filter_available)


@app.post("/jobs/filter-available/run")
def filter_available_job_sync(request: Request):
    try:
        return login_and_filter_available(base_url_from_request(request))
    except Exception as exc:
        return JSONResponse(
            status_code=500,
            content={"ok": False, "error": f"{type(exc).__name__}: {exc}"},
        )


@app.post("/jobs/open-availability-filter")
def open_availability_filter_job(request: Request):
    try:
        return login_and_open_availability_filter(base_url_from_request(request))
    except Exception as exc:
        return JSONResponse(
            status_code=500,
            content={"ok": False, "error": f"{type(exc).__name__}: {exc}"},
        )


@app.post("/jobs/open-availability-filter/start")
def open_availability_filter_job_start(request: Request):
    return start_background_job(
        request,
        "open-availability-filter",
        login_and_open_availability_filter,
    )


@app.post("/jobs/open-availability-filter/run")
def open_availability_filter_job_sync(request: Request):
    try:
        return login_and_open_availability_filter(base_url_from_request(request))
    except Exception as exc:
        return JSONResponse(
            status_code=500,
            content={"ok": False, "error": f"{type(exc).__name__}: {exc}"},
        )


@app.post("/jobs/scrape-products")
def scrape_products_job(request: Request):
    try:
        return scrape_all_available_products(base_url_from_request(request))
    except Exception as exc:
        return JSONResponse(
            status_code=500,
            content={"ok": False, "error": f"{type(exc).__name__}: {exc}"},
        )


@app.post("/jobs/scrape-products/start")
def scrape_products_job_start(request: Request):
    return start_background_job(request, "scrape-products", scrape_all_available_products)


@app.post("/jobs/scrape-products/run")
def scrape_products_job_sync(request: Request):
    try:
        return scrape_all_available_products(base_url_from_request(request))
    except Exception as exc:
        return JSONResponse(
            status_code=500,
            content={"ok": False, "error": f"{type(exc).__name__}: {exc}"},
        )


@app.get("/jobs")
def list_jobs():
    return {"ok": True, "jobs": list_job_snapshots()}


@app.get("/jobs/{job_id}", name="get_job")
def get_job(job_id: str):
    job = get_job_snapshot(job_id)

    if not job:
        return JSONResponse(status_code=404, content={"ok": False, "error": "job_not_found"})

    return job


@app.get("/screenshots/latest", name="get_latest_screenshot")
def get_latest_screenshot():
    screenshot_path = latest_screenshot_path()

    if not screenshot_path:
        return JSONResponse(status_code=404, content={"ok": False, "error": "no_screenshot_found"})

    return FileResponse(
        screenshot_path,
        media_type="image/png",
        filename=os.path.basename(screenshot_path),
    )


@app.get("/screenshots/{filename}", name="get_screenshot")
def get_screenshot(filename: str):
    try:
        safe_filename = safe_artifact_filename(filename)
    except RuntimeError as exc:
        return JSONResponse(status_code=400, content={"ok": False, "error": str(exc)})

    screenshot_path = os.path.join(ARTIFACTS_DIR, safe_filename)

    if not os.path.exists(screenshot_path):
        return JSONResponse(status_code=404, content={"ok": False, "error": "screenshot_not_found"})

    return FileResponse(
        screenshot_path,
        media_type="image/png",
        filename=safe_filename,
    )


@app.get("/reports/latest", name="get_latest_report")
def get_latest_report():
    report_path = latest_report_path()

    if not report_path:
        return JSONResponse(status_code=404, content={"ok": False, "error": "no_report_found"})

    return FileResponse(
        report_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=os.path.basename(report_path),
    )


@app.get("/reports/{filename}", name="get_report")
def get_report(filename: str):
    try:
        safe_filename = safe_report_filename(filename)
    except RuntimeError as exc:
        return JSONResponse(status_code=400, content={"ok": False, "error": str(exc)})

    report_path = os.path.join(ARTIFACTS_DIR, safe_filename)

    if not os.path.exists(report_path):
        return JSONResponse(status_code=404, content={"ok": False, "error": "report_not_found"})

    return FileResponse(
        report_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=safe_filename,
    )


@app.api_route("/{path_name:path}", methods=["GET", "POST", "PUT", "PATCH", "DELETE"])
def not_found(path_name):
    return JSONResponse(status_code=404, content={"ok": False, "error": "not_found"})


def main():
    import uvicorn

    port = int(os.environ.get("PORT", "8000"))
    uvicorn.run(app, host="0.0.0.0", port=port)


if __name__ == "__main__":
    main()
