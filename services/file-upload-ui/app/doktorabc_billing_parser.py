from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta, timezone, tzinfo
from io import BytesIO
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

from openpyxl import load_workbook


REQUIRED_COLUMNS = {
    "hash_id",
    "sent_date",
    "stock",
    "type",
    "total_medication_cost_incl_vat",
    "supply_price_base",
    "additional_cost",
    "waybill_id",
}
OPTIONAL_COLUMNS = {"uber_shipping_fee"}
HEADER_ALIASES = {
    "hash id": "hash_id",
    "sent date": "sent_date",
    "stock": "stock",
    "type": "type",
    "total medication cost incl. vat, €": "total_medication_cost_incl_vat",
    "total medication cost incl. vat, eur": "total_medication_cost_incl_vat",
    "total medication cost incl. vat": "total_medication_cost_incl_vat",
    "supply price base": "supply_price_base",
    "additional cost": "additional_cost",
    "waybill id": "waybill_id",
    "uber shipping fee": "uber_shipping_fee",
}
ALLOWED_TYPES = {"shipping", "RETURN: on_shelve", "reshipping"}
FOOTER_TYPE = "total medication cost incl. vat:"
DATE_TIME_FORMATS = (
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d %H:%M",
    "%Y-%m-%d",
    "%m/%d/%Y %H:%M:%S",
    "%m/%d/%Y %H:%M",
    "%m/%d/%Y",
    "%d.%m.%Y %H:%M:%S",
    "%d.%m.%Y %H:%M",
    "%d.%m.%Y",
)


@dataclass(frozen=True)
class BillingParseResult:
    rows: list[dict[str, object]]
    source_row_numbers: list[int]
    skipped_blank_rows: int
    skipped_footer_rows: int
    type_counts: dict[str, int]

    @property
    def skipped_rows(self) -> int:
        return self.skipped_blank_rows + self.skipped_footer_rows


def parse_doktorabc_billing_excel(
    contents: bytes,
    extension: str,
    billing_period_from: date,
    billing_period_to: date,
    timezone_name: str,
) -> BillingParseResult:
    rows = read_excel_rows(contents, extension)
    header_index, column_indexes = find_billing_header_row(rows)
    timezone = resolve_timezone(timezone_name)
    parsed_rows: list[dict[str, object]] = []
    source_row_numbers: list[int] = []
    type_counts: dict[str, int] = {}
    skipped_blank_rows = 0
    skipped_footer_rows = 0

    for row_offset, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        if is_blank_row(row):
            skipped_blank_rows += 1
            continue

        raw_type = cell_value(row, column_indexes, "type")
        type_text = normalize_text(raw_type)
        if type_text and normalize_header(type_text) == FOOTER_TYPE:
            skipped_footer_rows += 1
            continue

        hash_id = normalize_identifier(cell_value(row, column_indexes, "hash_id"))
        sent_date = parse_sent_date(cell_value(row, column_indexes, "sent_date"), timezone, row_offset)
        stock = normalize_stock(cell_value(row, column_indexes, "stock"))
        movement_type = normalize_text(raw_type)
        waybill_id = normalize_identifier(cell_value(row, column_indexes, "waybill_id"))

        missing = []
        if not hash_id:
            missing.append("Hash id")
        if not sent_date:
            missing.append("Sent date")
        if not stock:
            missing.append("Stock")
        if not movement_type:
            missing.append("Type")
        if missing:
            raise ValueError(f"Row {row_offset}: missing required value(s): {', '.join(missing)}.")

        if movement_type not in ALLOWED_TYPES:
            allowed = ", ".join(sorted(ALLOWED_TYPES))
            raise ValueError(f"Row {row_offset}: unsupported Type {movement_type!r}. Expected one of: {allowed}.")

        total_cost = parse_numeric(
            cell_value(row, column_indexes, "total_medication_cost_incl_vat"),
            "Total medication cost incl. VAT",
            row_offset,
        )
        supply_price_base = parse_numeric(
            cell_value(row, column_indexes, "supply_price_base"),
            "Supply Price Base",
            row_offset,
        )
        additional_cost = parse_numeric(
            cell_value(row, column_indexes, "additional_cost"),
            "Additional Cost",
            row_offset,
        )
        uber_shipping_fee = parse_numeric(
            cell_value(row, column_indexes, "uber_shipping_fee"),
            "Uber Shipping Fee",
            row_offset,
            default=0,
        )

        type_counts[movement_type] = type_counts.get(movement_type, 0) + 1
        source_row_numbers.append(row_offset)
        parsed_rows.append(
            {
                "hash_id": hash_id,
                "sent_date": sent_date,
                "stock": stock,
                "type": movement_type,
                "total_medication_cost_incl_vat": total_cost,
                "supply_price_base": supply_price_base,
                "additional_cost": additional_cost,
                "waybill_id": waybill_id,
                "uber_shipping_fee": uber_shipping_fee,
                "billing_period_from": billing_period_from.isoformat(),
                "billing_period_to": billing_period_to.isoformat(),
            }
        )

    if not parsed_rows:
        raise ValueError("No valid DoktorABC Abrechnung rows found in the Excel file.")

    return BillingParseResult(
        rows=parsed_rows,
        source_row_numbers=source_row_numbers,
        skipped_blank_rows=skipped_blank_rows,
        skipped_footer_rows=skipped_footer_rows,
        type_counts=type_counts,
    )


def read_excel_rows(contents: bytes, extension: str) -> list[list[object]]:
    extension = extension.lower()
    if extension == ".xlsx":
        return read_xlsx_rows(contents)
    if extension == ".xls":
        return read_xls_rows(contents)
    raise ValueError("Only .xlsx and .xls files can be parsed for DoktorABC Abrechnung.")


def read_xlsx_rows(contents: bytes) -> list[list[object]]:
    workbook = load_workbook(BytesIO(contents), read_only=True, data_only=True)
    try:
        sheet = workbook.active
        return [list(row) for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def read_xls_rows(contents: bytes) -> list[list[object]]:
    from xlrd import XL_CELL_DATE, open_workbook, xldate_as_datetime

    workbook = open_workbook(file_contents=contents)
    sheet = workbook.sheet_by_index(0)
    rows = []

    for row_index in range(sheet.nrows):
        values = []
        for col_index in range(sheet.ncols):
            cell = sheet.cell(row_index, col_index)
            if cell.ctype == XL_CELL_DATE:
                values.append(xldate_as_datetime(cell.value, workbook.datemode))
            else:
                values.append(cell.value)
        rows.append(values)

    return rows


def find_billing_header_row(rows: list[list[object]]) -> tuple[int, dict[str, int]]:
    for row_index, row in enumerate(rows[:20]):
        column_indexes = {}
        for col_index, value in enumerate(row):
            column = HEADER_ALIASES.get(normalize_header(value))
            if column and column not in column_indexes:
                column_indexes[column] = col_index

        if REQUIRED_COLUMNS.issubset(column_indexes):
            return row_index, column_indexes

    expected = ", ".join(sorted(REQUIRED_COLUMNS | OPTIONAL_COLUMNS))
    raise ValueError(f"Could not find the DoktorABC Abrechnung header row. Expected columns: {expected}.")


def cell_value(row: list[object], column_indexes: dict[str, int], column: str) -> object | None:
    col_index = column_indexes.get(column)
    if col_index is None or col_index >= len(row):
        return None
    return row[col_index]


def normalize_header(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\ufeff", "").strip()).lower()


def is_blank_row(row: list[object]) -> bool:
    return all(normalize_text(value) is None for value in row)


def normalize_text(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        text = value.strftime("%Y-%m-%d %H:%M:%S")
    elif isinstance(value, date):
        text = value.isoformat()
    elif isinstance(value, float) and value.is_integer():
        text = str(int(value))
    else:
        text = str(value)

    text = text.replace("\r\n", "\n").replace("\r", "\n").strip()
    if not text:
        return None

    return re.sub(r"\s+", " ", text).strip()


def normalize_stock(value: object) -> str | None:
    text = None if value is None else str(value).replace("\r\n", "\n").replace("\r", "\n").strip()
    if not text:
        return None

    parts = [re.sub(r"\s+", " ", part).strip() for part in text.split("\n") if part.strip()]
    return " | ".join(parts) if parts else None


def normalize_identifier(value: object) -> str | None:
    text = normalize_text(value)
    return text if text else None


def resolve_timezone(timezone_name: str) -> tzinfo:
    try:
        return ZoneInfo(timezone_name or "Europe/Berlin")
    except ZoneInfoNotFoundError:
        try:
            return ZoneInfo("Europe/Berlin")
        except ZoneInfoNotFoundError:
            return timezone(timedelta(hours=1))


def parse_sent_date(value: object, timezone: tzinfo, row_number: int) -> str | None:
    if value is None:
        return None

    if isinstance(value, datetime):
        parsed = value
    elif isinstance(value, date):
        parsed = datetime.combine(value, time.min)
    else:
        text = normalize_text(value)
        if not text:
            return None
        parsed = parse_datetime_text(text, row_number)

    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone)

    return parsed.isoformat()


def parse_datetime_text(value: str, row_number: int) -> datetime:
    for date_format in DATE_TIME_FORMATS:
        try:
            return datetime.strptime(value, date_format)
        except ValueError:
            continue

    raise ValueError(f"Row {row_number}: could not parse Sent date {value!r}.")


def parse_numeric(value: object, column_name: str, row_number: int, default: float | None = None) -> float:
    if value is None or normalize_text(value) is None:
        if default is not None:
            return default
        raise ValueError(f"Row {row_number}: missing numeric value for {column_name}.")

    if isinstance(value, int | float):
        return round(float(value), 2)

    text = str(value).replace("\xa0", "").replace("€", "").strip()
    text = re.sub(r"[^0-9,.\-]", "", text)
    if not text:
        if default is not None:
            return default
        raise ValueError(f"Row {row_number}: missing numeric value for {column_name}.")

    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")

    try:
        return round(float(text), 2)
    except ValueError as exc:
        raise ValueError(f"Row {row_number}: invalid numeric value for {column_name}: {value!r}.") from exc
