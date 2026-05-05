from __future__ import annotations

import re
from datetime import date, datetime
from io import BytesIO

from openpyxl import load_workbook
from xlrd import XL_CELL_DATE, open_workbook, xldate_as_datetime


ORDER_COLUMNS = [
    "Order id",
    "Customer uuid",
    "Customer name",
    "Customer birthday",
    "Customer address",
    "Products",
    "Tracking ID",
    "Tracking URL",
    "PZNs",
    "Doctor",
    "Doctor address",
    "Prescription date",
    "Sent date",
    "Returns",
]

DATE_ONLY_COLUMNS = {"Customer birthday"}
DATE_TIME_COLUMNS = {"Prescription date", "Sent date", "Returns"}
DATE_PATTERN = re.compile(
    r"^\s*(\d{1,2})[./-](\d{1,2})[./-](\d{2,4})"
    r"(?:\s+(\d{1,2}):(\d{2})(?::(\d{2}))?)?\s*$"
)


def parse_oed_orders_excel(contents: bytes, extension: str) -> list[dict[str, str | None]]:
    rows = read_excel_rows(contents, extension)
    header_index, column_indexes = find_header_row(rows)
    orders = []

    for row in rows[header_index + 1 :]:
        order = {}
        for column in ORDER_COLUMNS:
            value = row[column_indexes[column]] if column_indexes[column] < len(row) else None
            order[column] = normalize_column_value(column, value)

        if any(value is not None for value in order.values()):
            orders.append(order)

    if not orders:
        raise ValueError("No OED order rows found in the Excel file.")

    return orders


def read_excel_rows(contents: bytes, extension: str) -> list[list[object]]:
    extension = extension.lower()
    if extension == ".xlsx":
        return read_xlsx_rows(contents)
    if extension == ".xls":
        return read_xls_rows(contents)
    raise ValueError("Only .xlsx and .xls files can be parsed for OED orders.")


def read_xlsx_rows(contents: bytes) -> list[list[object]]:
    workbook = load_workbook(BytesIO(contents), read_only=True, data_only=True)
    try:
        sheet = workbook.active
        return [list(row) for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def read_xls_rows(contents: bytes) -> list[list[object]]:
    workbook = open_workbook(file_contents=contents)
    sheet = workbook.sheet_by_index(0)
    rows = []

    for row_index in range(sheet.nrows):
        values = []
        for col_index in range(sheet.ncols):
            cell = sheet.cell(row_index, col_index)
            if cell.ctype == XL_CELL_DATE:
                values.append(xldate_as_datetime(cell.value, workbook.datemode))
            else:
                values.append(cell.value)
        rows.append(values)

    return rows


def find_header_row(rows: list[list[object]]) -> tuple[int, dict[str, int]]:
    expected_headers = {normalize_header(column): column for column in ORDER_COLUMNS}

    for row_index, row in enumerate(rows[:20]):
        column_indexes = {}
        for col_index, value in enumerate(row):
            column = expected_headers.get(normalize_header(value))
            if column and column not in column_indexes:
                column_indexes[column] = col_index

        if len(column_indexes) == len(ORDER_COLUMNS):
            return row_index, column_indexes

    missing = ", ".join(ORDER_COLUMNS)
    raise ValueError(f"Could not find the OED header row. Expected columns: {missing}.")


def normalize_header(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\ufeff", "").strip()).lower()


def normalize_column_value(column: str, value: object) -> str | None:
    if column in DATE_ONLY_COLUMNS:
        return format_excel_date(value, include_time=False)
    if column in DATE_TIME_COLUMNS:
        return format_excel_date(value, include_time=True)
    return normalize_text(value)


def format_excel_date(value: object, include_time: bool) -> str | None:
    if value is None:
        return None

    if isinstance(value, datetime):
        return format_datetime(value, include_time)

    if isinstance(value, date):
        return value.strftime("%m/%d/%Y")

    text = normalize_text(value)
    if text is None:
        return None

    match = DATE_PATTERN.match(text)
    if not match:
        return text

    day, month, year, hour, minute, second = match.groups()
    year = normalize_year(year)
    formatted = f"{int(month):02d}/{int(day):02d}/{year:04d}"
    if include_time and hour is not None:
        formatted += f" {int(hour):02d}:{int(minute):02d}:{int(second or 0):02d}"
    return formatted


def format_datetime(value: datetime, include_time: bool) -> str:
    if include_time and (value.hour or value.minute or value.second):
        return value.strftime("%m/%d/%Y %H:%M:%S")
    return value.strftime("%m/%d/%Y")


def normalize_year(year: str) -> int:
    number = int(year)
    if number < 100:
        return 2000 + number if number < 70 else 1900 + number
    return number


def normalize_text(value: object) -> str | None:
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.strftime("%m/%d/%Y %H:%M:%S")

    if isinstance(value, date):
        return value.strftime("%m/%d/%Y")

    if isinstance(value, float) and value.is_integer():
        text = str(int(value))
    else:
        text = str(value)

    text = text.replace("\r\n", "\n").replace("\r", "\n").strip()
    if not text:
        return None

    parts = [collapse_spaces(part) for part in text.split("\n") if part.strip()]
    if not parts:
        return None

    return " | ".join(parts)


def collapse_spaces(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()
